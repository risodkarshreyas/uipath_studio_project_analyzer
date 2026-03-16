#!/usr/bin/env python3
"""
UiPath RPA Project Analyzer  v2.0
===================================
Scans a single project OR an entire repository of UiPath Studio projects.
Generates a full Activity/Application report PLUS a Reusability Matrix with:
  - Usage frequency across all projects
  - Rule-based description of what each workflow/activity solves
  - Reusability score (0–100) per workflow and per activity category
  - Shared library candidate recommendations

Usage:
  # Single project
  python3 uipath_analyzer.py ./MyProject --output all

  # Entire repo (auto-discovers all projects)
  python3 uipath_analyzer.py ./MyRepo --repo --output all --report-dir ./reports

  # Repo, console only
  python3 uipath_analyzer.py ./MyRepo --repo
"""

import os, sys, json, argparse, re
from pathlib import Path
from datetime import datetime
from xml.etree import ElementTree as ET
from collections import defaultdict
from copy import deepcopy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ══════════════════════════════════════════════════════════
#  ANSI colour helpers
# ══════════════════════════════════════════════════════════
class C:
    RESET="\033[0m"; BOLD="\033[1m"; CYAN="\033[96m"; GREEN="\033[92m"
    YELLOW="\033[93m"; RED="\033[91m"; BLUE="\033[94m"; MAGENTA="\033[95m"
    WHITE="\033[97m"; GREY="\033[90m"

def banner():
    print(f"""{C.CYAN}{C.BOLD}
 ██╗   ██╗██╗██████╗  █████╗ ████████╗██╗  ██╗
 ██║   ██║██║██╔══██╗██╔══██╗╚══██╔══╝██║  ██║
 ██║   ██║██║██████╔╝███████║   ██║   ███████║
 ██║   ██║██║██╔═══╝ ██╔══██║   ██║   ██╔══██║
 ╚██████╔╝██║██║     ██║  ██║   ██║   ██║  ██║
  ╚═════╝ ╚═╝╚═╝     ╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝
{C.RESET}{C.BLUE}  UiPath RPA Analyzer + Reusability Matrix  |  v2.1{C.RESET}
""")

# ══════════════════════════════════════════════════════════
#  Activity Knowledge Base
# ══════════════════════════════════════════════════════════
ACTIVITY_CATALOG = {
    "ExcelApplicationScope":("Excel","Excel Application Scope"),
    "ReadRange":("Excel","Read Range"), "WriteRange":("Excel","Write Range"),
    "WriteCell":("Excel","Write Cell"), "ReadCell":("Excel","Read Cell"),
    "AppendRange":("Excel","Append Range"), "SaveWorkbook":("Excel","Save Workbook"),
    "CloseWorkbook":("Excel","Close Workbook"), "InsertRows":("Excel","Insert Rows"),
    "DeleteRows":("Excel","Delete Rows"), "SortRange":("Excel","Sort Range"),
    "FilterTable":("Excel","Filter Table"), "GetWorkbookSheets":("Excel","Get Workbook Sheets"),
    "OpenBrowser":("Web Browser","Open Browser"), "CloseBrowser":("Web Browser","Close Browser"),
    "Navigate":("Web Browser","Navigate To"), "NavigateTo":("Web Browser","Navigate To"),
    "UploadFile":("Web Browser","Upload File"), "GetText":("Web Browser","Get Text"),
    "InjectJsScript":("Web Browser","Inject JS Script"), "SelectItem":("Web Browser","Select Item"),
    "WaitForPageLoad":("Web Browser","Wait For Page Load"),
    "Click":("UI Automation","Click"), "TypeInto":("UI Automation","Type Into"),
    "SendHotkey":("UI Automation","Send Hotkey"), "AttachWindow":("UI Automation","Attach Window"),
    "AttachBrowser":("UI Automation","Attach Browser"), "FindElement":("UI Automation","Find Element"),
    "WaitElementVanish":("UI Automation","Wait Element Vanish"),
    "DoubleClick":("UI Automation","Double Click"), "RightClick":("UI Automation","Right Click"),
    "Hover":("UI Automation","Hover"), "SetText":("UI Automation","Set Text"),
    "TakeScreenshot":("UI Automation","Take Screenshot"),
    "GetAttribute":("UI Automation","Get Attribute"),
    "GetOutlookMailMessages":("Outlook/Email","Get Outlook Mail Messages"),
    "SendOutlookMailMessage":("Outlook/Email","Send Outlook Mail Message"),
    "SaveMailMessage":("Outlook/Email","Save Mail Message"),
    "MoveOutlookMailMessage":("Outlook/Email","Move Outlook Mail Message"),
    "MarkAsRead":("Outlook/Email","Mark Mail As Read"),
    "DeleteMailMessage":("Outlook/Email","Delete Mail Message"),
    "SendSmtpMailMessage":("Outlook/Email","Send SMTP Mail Message"),
    "GetImapMailMessages":("Outlook/Email","Get IMAP Mail Messages"),
    "ReadPDFText":("PDF","Read PDF Text"), "ReadPDFWithOCR":("PDF","Read PDF With OCR"),
    "GetPDFPageCount":("PDF","Get PDF Page Count"),
    "JoinPDFFiles":("PDF","Join PDF Files"), "ExtractPDFPage":("PDF","Extract PDF Page"),
    "PresentValidationStation":("Document Understanding","Validation Station"),
    "IntelligentOCRActivity":("Document Understanding","Digitize Document"),
    "ClassifyDocumentScope":("Document Understanding","Classify Document"),
    "ExtractDocumentData":("Document Understanding","Extract Document Data"),
    "SetTransactionStatus":("Orchestrator","Set Transaction Status"),
    "GetTransactionItem":("Orchestrator","Get Transaction Item"),
    "AddQueueItem":("Orchestrator","Add Queue Item"),
    "BulkAddQueueItems":("Orchestrator","Bulk Add Queue Items"),
    "GetQueueItems":("Orchestrator","Get Queue Items"),
    "GetAssetActivity":("Orchestrator","Get Asset"),
    "GetCredentialActivity":("Orchestrator","Get Credential"),
    "SetAssetActivity":("Orchestrator","Set Asset"),
    "LogMessage":("System","Log Message"), "AssignActivity":("System","Assign"),
    "InvokeWorkflowFile":("System","Invoke Workflow File"),
    "InvokeCode":("System","Invoke Code"), "InvokePowerShell":("System","Invoke PowerShell"),
    "StartProcess":("System","Start Process"), "KillProcess":("System","Kill Process"),
    "ReadTextFile":("System","Read Text File"), "WriteTextFile":("System","Write Text File"),
    "AppendLine":("System","Append Line"), "MoveFile":("System","Move File"),
    "CopyFile":("System","Copy File"), "DeleteFile":("System","Delete File"),
    "CompressZipFiles":("System","Compress Zip Files"),
    "HttpClientActivity":("HTTP","HTTP Request"),
    "Connect":("Database","Connect"), "Disconnect":("Database","Disconnect"),
    "ExecuteQuery":("Database","Execute Query"),
    "ExecuteNonQuery":("Database","Execute Non Query"),
    "InsertDataTable":("Database","Insert DataTable"),
    "ForEach":("Control Flow","For Each"), "ForEachRow":("Control Flow","For Each Row"),
    "While":("Control Flow","While"), "If":("Control Flow","If"),
    "Switch":("Control Flow","Switch"), "TryCatch":("Control Flow","Try Catch"),
    "Throw":("Control Flow","Throw"), "Sequence":("Control Flow","Sequence"),
    "Flowchart":("Control Flow","Flowchart"),
}

APP_MAP = {
    "Excel":"Microsoft Excel", "Web Browser":"Web Browser (Chrome/Edge/Firefox)",
    "Outlook/Email":"Microsoft Outlook / Email", "PDF":"PDF Files",
    "Document Understanding":"UiPath Document Understanding / IDP",
    "Orchestrator":"UiPath Orchestrator", "SAP":"SAP ERP",
    "Microsoft Word":"Microsoft Word", "Salesforce":"Salesforce CRM",
    "Database":"Database (SQL/OLEDB)", "HTTP":"HTTP / REST APIs",
    "UI Automation":"Windows Desktop Application",
}

WINDOWS_APP_PATTERNS = {
    r"sap":"SAP ERP (Windows GUI)", r"outlook":"Microsoft Outlook",
    r"excel":"Microsoft Excel", r"winword":"Microsoft Word",
    r"chrome":"Google Chrome", r"firefox":"Mozilla Firefox",
    r"msedge|edge":"Microsoft Edge", r"acrobat":"Adobe Acrobat",
    r"salesforce":"Salesforce (Browser)", r"servicenow":"ServiceNow",
    r"workday":"Workday",
}

# ══════════════════════════════════════════════════════════
#  Reusability Knowledge Base  (rule-based, fully offline)
# ══════════════════════════════════════════════════════════

# Per-category: base reusability score (0-100) + description of value
CATEGORY_REUSABILITY = {
    "Orchestrator":      (92, "Queue & transaction management; forms the RE Framework backbone reusable across any automation."),
    "System":            (88, "Core utilities (logging, file ops, workflow invocation) applicable to virtually every RPA project."),
    "Outlook/Email":     (85, "Email notification & retrieval patterns that repeat in most business process automations."),
    "Excel":             (82, "Spreadsheet read/write operations commonly shared across Finance, HR and Operations workflows."),
    "PDF":               (78, "Document extraction logic reusable wherever PDF invoices, receipts or reports are processed."),
    "Document Understanding": (75, "IDP extraction pipeline reusable across invoice, PO, and contract processing projects."),
    "Database":          (80, "Database connectivity & query patterns shareable across all data-intensive automations."),
    "HTTP":              (77, "REST API integration wrappers reusable across any project consuming web services."),
    "Web Browser":       (60, "Browser automation steps — reusable when targeting the same web application."),
    "UI Automation":     (50, "Screen interaction steps — lower reusability as they are selector-dependent per application."),
    "Control Flow":      (70, "Standard error handling & loop patterns (TryCatch, ForEach) reusable as framework templates."),
    "SAP":               (65, "SAP GUI automation steps — reusable within SAP-focused project families."),
    "Microsoft Word":    (72, "Word document generation patterns reusable in reporting workflows."),
    "Salesforce":        (68, "Salesforce CRM interaction patterns reusable across Sales/Support automations."),
}

# Workflow-name pattern → (problem it solves, reusability boost 0-20)
WORKFLOW_PATTERNS = [
    (r"send.*email|email.*notif|notification",  "Sends automated email notifications on process completion or failure.", 20),
    (r"error.*handl|exception.*handl|handle.*error", "Centralised exception handling and alerting logic.", 18),
    (r"login|authenticate|credential",          "Handles application login and credential retrieval.", 17),
    (r"read.*excel|excel.*read|load.*data",     "Reads and loads structured data from Excel spreadsheets.", 15),
    (r"write.*excel|excel.*write|save.*data",   "Writes or updates processed data back to Excel.", 14),
    (r"extract.*pdf|pdf.*extract|ocr",          "Extracts text or structured data from PDF documents using OCR.", 16),
    (r"validate|validation",                    "Validates extracted or input data against business rules.", 15),
    (r"queue|transaction",                      "Manages Orchestrator queue items and transaction lifecycle.", 18),
    (r"log|logging",                            "Provides centralised logging and audit trail.", 16),
    (r"report|summary",                         "Generates and distributes process summary reports.", 14),
    (r"sap.*post|post.*sap|sap.*entry",         "Posts data entries into SAP ERP modules.", 13),
    (r"ad.*account|active.*directory",          "Creates or manages Active Directory user accounts.", 14),
    (r"db.*connect|database.*connect|connect",  "Establishes and manages database connections.", 15),
    (r"download|fetch.*file|get.*file",         "Downloads or retrieves files from remote sources.", 13),
    (r"archive|move.*file|cleanup",             "Archives processed files and performs cleanup operations.", 12),
    (r"config|setting|parameter",               "Loads runtime configuration and parameters.", 17),
    (r"main|orchestrat",                        "Orchestrates the end-to-end process flow calling sub-workflows.", 10),
]

# Activity tag patterns → problem description
ACTIVITY_DESCRIPTIONS = {
    "Orchestrator":  "Manages process queues, credentials and transaction state via UiPath Orchestrator.",
    "System":        "Handles file operations, logging, and invocation of modular sub-workflows.",
    "Outlook/Email": "Automates email retrieval, dispatch and inbox management via Microsoft Outlook.",
    "Excel":         "Reads, writes and manipulates data in Microsoft Excel workbooks.",
    "PDF":           "Extracts and processes text and structured data from PDF documents.",
    "Document Understanding": "Applies intelligent document processing (IDP) to classify and extract unstructured data.",
    "Database":      "Connects to and queries relational databases for data retrieval or updates.",
    "HTTP":          "Integrates with external REST/SOAP APIs via HTTP requests.",
    "Web Browser":   "Automates interactions with web-based applications in the browser.",
    "UI Automation": "Drives desktop or web UI elements (clicks, typing, reading values).",
    "Control Flow":  "Provides structured error handling, loops and conditional branching.",
    "SAP":           "Automates SAP ERP GUI transactions and data entry.",
    "Microsoft Word":"Creates or manipulates Microsoft Word documents programmatically.",
    "Salesforce":    "Automates Salesforce CRM record management and queries.",
}

SHARED_LIB_THRESHOLD = 60   # score >= this → candidate for shared library


# ══════════════════════════════════════════════════════════
#  Single-Project Analyzer
# ══════════════════════════════════════════════════════════
class UiPathProjectAnalyzer:
    def __init__(self, project_path: Path):
        self.project_path = Path(project_path).resolve()
        self.project_json = {}
        self.xaml_files   = []
        self.results = {
            "project_info": {}, "activities": [],
            "applications": set(), "web_urls": set(),
            "window_selectors": set(), "invoked_workflows": [],
            "summary": {}, "unknown_activities": [],
            "workflow_details": [],          # per-XAML file breakdown
            # ── NEW: deep target extraction ──────────────────────
            "web_targets":     [],   # {url, page_title, browser, activity, xaml_file}
            "window_targets":  [],   # {app_exe, window_title, activity, xaml_file}
        }

    def discover(self):
        pj = self.project_path / "project.json"
        if not pj.exists():
            return False
        with open(pj, "r", encoding="utf-8") as f:
            self.project_json = json.load(f)
        self.xaml_files = sorted(self.project_path.rglob("*.xaml"))
        return True

    def parse_all_xaml(self):
        activity_map = defaultdict(lambda: {
            "count":0,"category":"","friendly_name":"","raw_tag":"","files":set()
        })
        for xf in self.xaml_files:
            rel = xf.relative_to(self.project_path)
            self._parse_xaml(xf, rel, activity_map)
        for raw_tag, info in activity_map.items():
            info["files"] = sorted(info["files"])
            info["raw_tag"] = raw_tag
            self.results["activities"].append(dict(info))
        self.results["activities"].sort(key=lambda x:(x["category"],-x["count"]))

    def _parse_xaml(self, xaml_path, rel_path, activity_map):
        try:
            tree = ET.parse(xaml_path)
            root = tree.getroot()
        except ET.ParseError:
            return
        # Per-workflow detail
        wf_detail = {
            "file": str(rel_path), "display_name": root.get("DisplayName",""),
            "activities": defaultdict(int), "categories": set(),
            "invokes": []
        }
        self._walk_element(root, str(rel_path), activity_map, wf_detail)
        wf_detail["activities"] = dict(wf_detail["activities"])
        wf_detail["categories"] = sorted(wf_detail["categories"])
        self.results["workflow_details"].append(wf_detail)

    def _local_name(self, tag):
        return tag.split("}")[-1] if "}" in tag else tag

    def _walk_element(self, elem, filename, activity_map, wf_detail):
        local = self._local_name(elem.tag)
        if local not in ("Activity","Members","x:Members"):
            cat, friendly = self._classify_activity(local, elem)
            if cat:
                activity_map[local]["count"]        += 1
                activity_map[local]["category"]      = cat
                activity_map[local]["friendly_name"] = friendly
                activity_map[local]["files"].add(filename)
                wf_detail["activities"][local] = wf_detail["activities"].get(local,0)+1
                wf_detail["categories"].add(cat)
                app = APP_MAP.get(cat)
                if app:
                    self.results["applications"].add(app)

                display_name = elem.get("DisplayName", local)

                # ── URL extraction (direct attributes) ────────────
                for url_attr in ("Url","url","URL","UrlString"):
                    url = elem.get(url_attr, "")
                    if url and url.startswith("http"):
                        self.results["web_urls"].add(url)
                        self._record_web_target(
                            url=url, page_title="", browser=self._guess_browser(elem),
                            activity=display_name, xaml_file=filename
                        )

                # ── Selector deep-parse ───────────────────────────
                sel_raw = elem.get("Selector") or elem.get("selector") or ""
                if sel_raw:
                    self._deep_parse_selector(sel_raw, display_name, filename)

                # ── InvokeWorkflowFile ────────────────────────────
                if local == "InvokeWorkflowFile":
                    wf = elem.get("InvokedWorkflowFileName") or ""
                    if wf:
                        self.results["invoked_workflows"].append(
                            {"workflow":wf,"called_from":filename})
                        wf_detail["invokes"].append(wf)

        for child in elem:
            self._walk_element(child, filename, activity_map, wf_detail)

    # ── Deep selector parser ───────────────────────────────────────
    def _deep_parse_selector(self, selector_raw: str, activity: str, xaml_file: str):
        """
        Parse UiPath selector XML (possibly HTML-entity-encoded) and extract:
          - <html app='...' title='...' url='...'>  → web targets
          - <wnd app='...' title='...'>             → window targets
          - <ctrl ...>, <webctrl ...>               → supplemental info
        """
        # Decode HTML entities that UiPath uses in selector attributes
        sel = selector_raw.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&").replace("&quot;",'"')

        # Wrap in a root so ET can parse it
        try:
            root = ET.fromstring(f"<root>{sel}</root>")
        except ET.ParseError:
            # Fallback: regex scan
            self._regex_selector_fallback(sel, activity, xaml_file)
            return

        app_exe    = ""
        window_title = ""
        browser    = ""
        page_title = ""
        url        = ""

        for node in root:
            tag = node.tag.lower()
            attrs = {k.lower(): v for k, v in node.attrib.items()}

            if tag == "wnd":
                a = attrs.get("app","")
                t = attrs.get("title","")
                if a: app_exe = a
                if t: window_title = t

            elif tag in ("html","browser","web"):
                a = attrs.get("app","")
                t = attrs.get("title","")
                u = attrs.get("url","") or attrs.get("urlpart","")
                if a: browser = a
                if t: page_title = t
                if u and u.startswith("http"): url = u

            elif tag == "webctrl":
                # webctrl can sometimes carry a url
                u = attrs.get("url","") or attrs.get("urlpart","")
                if u and u.startswith("http") and not url:
                    url = u

        # Record web target if we found browser/page info
        if browser or page_title or url:
            self._record_web_target(url, page_title, browser, activity, xaml_file)

        # Record window target if we found wnd info
        if app_exe or window_title:
            self._record_window_target(app_exe, window_title, activity, xaml_file)

    def _regex_selector_fallback(self, sel: str, activity: str, xaml_file: str):
        """Regex-based selector scan when XML parse fails."""
        # Window app/title
        wnd_match = re.search(r"<wnd[^>]*app='([^']*)'[^>]*title='([^']*)'", sel, re.I)
        if not wnd_match:
            wnd_match = re.search(r"<wnd[^>]*title='([^']*)'[^>]*app='([^']*)'", sel, re.I)
        if wnd_match:
            self._record_window_target(wnd_match.group(1), wnd_match.group(2), activity, xaml_file)

        # Browser/HTML
        html_match = re.search(r"<html[^>]*app='([^']*)'[^>]*(?:title='([^']*)')?[^>]*(?:url='([^']*)')?", sel, re.I)
        if html_match:
            self._record_web_target(
                html_match.group(3) or "", html_match.group(2) or "",
                html_match.group(1) or "", activity, xaml_file
            )

        # Loose URL scan
        for url in re.findall(r"https?://[^\s'\"<>]+", sel):
            self._record_web_target(url, "", "", activity, xaml_file)

    def _guess_browser(self, elem) -> str:
        """Guess browser from BrowserType attribute."""
        bt = elem.get("BrowserType","") or elem.get("browserType","")
        mapping = {"Chrome":"chrome.exe","Firefox":"firefox.exe",
                   "Edge":"msedge.exe","IE":"iexplore.exe","InternetExplorer":"iexplore.exe"}
        return mapping.get(bt, bt)

    def _record_web_target(self, url, page_title, browser, activity, xaml_file):
        """Deduplicated append to web_targets."""
        # Normalise browser name
        browser_map = {
            "chrome.exe":"Google Chrome","firefox.exe":"Mozilla Firefox",
            "msedge.exe":"Microsoft Edge","iexplore.exe":"Internet Explorer",
        }
        browser_friendly = browser_map.get(browser.lower(), browser) if browser else ""

        key = (url.rstrip("/"), page_title, browser_friendly)
        # Avoid duplicate (url+title+browser) combos
        existing_keys = {
            (t["url"].rstrip("/"), t["page_title"], t["browser"])
            for t in self.results["web_targets"]
        }
        if key not in existing_keys and (url or page_title):
            self.results["web_targets"].append({
                "url":          url,
                "page_title":   page_title,
                "browser":      browser_friendly,
                "activity":     activity,
                "xaml_file":    xaml_file,
            })
        if url and url.startswith("http"):
            self.results["web_urls"].add(url)

    def _record_window_target(self, app_exe, window_title, activity, xaml_file):
        """Deduplicated append to window_targets."""
        # Map known exe names to friendly app names
        exe_friendly_map = {
            "saplogon.exe":        "SAP Logon / SAP GUI",
            "payrolltool.exe":     "Payroll Management System",
            "financeapproval.exe": "Finance Approval System",
            "excel.exe":           "Microsoft Excel",
            "winword.exe":         "Microsoft Word",
            "powerpnt.exe":        "Microsoft PowerPoint",
            "outlook.exe":         "Microsoft Outlook",
            "notepad.exe":         "Notepad",
            "powershell.exe":      "PowerShell",
            "cmd.exe":             "Command Prompt",
            "acrobat.exe":         "Adobe Acrobat",
            "acrord32.exe":        "Adobe Acrobat Reader",
        }
        app_friendly = exe_friendly_map.get(app_exe.lower(), app_exe)

        # Also detect app from window title keywords
        if not app_friendly or app_friendly == app_exe:
            title_lower = window_title.lower()
            for pattern, friendly in WINDOWS_APP_PATTERNS.items():
                if re.search(pattern, title_lower) or re.search(pattern, app_exe.lower()):
                    app_friendly = friendly
                    break

        key = (app_exe.lower(), window_title)
        existing_keys = {
            (t["app_exe"].lower(), t["window_title"])
            for t in self.results["window_targets"]
        }
        if key not in existing_keys and (app_exe or window_title):
            self.results["window_targets"].append({
                "app_exe":       app_exe,
                "app_friendly":  app_friendly,
                "window_title":  window_title,
                "activity":      activity,
                "xaml_file":     xaml_file,
            })
        # Also register in applications set
        if app_friendly:
            self.results["applications"].add(app_friendly)
            self.results["window_selectors"].add(app_friendly)

    def _classify_activity(self, local_name, elem):
        if local_name in ACTIVITY_CATALOG:
            return ACTIVITY_CATALOG[local_name]
        lower = local_name.lower()
        if any(k in lower for k in ["excel","workbook","sheet","range","cell"]):
            return ("Excel", local_name)
        if any(k in lower for k in ["browser","web","navigate","url","http"]):
            return ("Web Browser", local_name)
        if any(k in lower for k in ["click","type","hotkey","attach","find","element","window"]):
            return ("UI Automation", local_name)
        if any(k in lower for k in ["mail","email","outlook","smtp","imap"]):
            return ("Outlook/Email", local_name)
        if any(k in lower for k in ["pdf","ocr"]):
            return ("PDF", local_name)
        if any(k in lower for k in ["queue","transaction","asset","credential"]):
            return ("Orchestrator", local_name)
        if any(k in lower for k in ["sap"]):
            return ("SAP", local_name)
        if any(k in lower for k in ["database","sql","query","executenonquery","connect"]):
            return ("Database", local_name)
        if any(k in lower for k in ["log","assign","invoke","delay","file","folder","process","zip"]):
            return ("System", local_name)
        return (None, None)

    def extract_project_info(self):
        pj = self.project_json
        self.results["project_info"] = {
            "name": pj.get("name","N/A"),
            "description": pj.get("description","N/A"),
            "main_workflow": pj.get("main","N/A"),
            "project_version": pj.get("projectVersion","N/A"),
            "dependencies": pj.get("dependencies",{}),
            "analyzed_on": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "path": str(self.project_path),
        }

    def build_summary(self):
        acts = self.results["activities"]
        self.results["summary"] = {
            "total_activities_unique": len(acts),
            "total_activity_usages": sum(a["count"] for a in acts),
            "total_xaml_files": len(self.xaml_files),
            "applications_detected": len(self.results["applications"]),
            "web_urls_found": len(self.results["web_urls"]),
            "web_targets_found": len(self.results["web_targets"]),
            "window_targets_found": len(self.results["window_targets"]),
            "invoked_workflows": len(self.results["invoked_workflows"]),
            "categories": sorted(set(a["category"] for a in acts)),
        }

    def run(self):
        if not self.discover():
            return False
        self.parse_all_xaml()
        self.extract_project_info()
        self.build_summary()
        return True


# ══════════════════════════════════════════════════════════
#  Reusability Engine  (rule-based, fully offline)
# ══════════════════════════════════════════════════════════
class ReusabilityEngine:
    def __init__(self, projects: list):
        """projects: list of UiPathProjectAnalyzer (already run)"""
        self.projects = projects

    # ── Workflow-level matrix ──────────────────────────────
    def workflow_matrix(self) -> list:
        """
        Returns one row per unique workflow filename (basename),
        scored and described.
        """
        # Aggregate: workflow_name → {projects, total_usages, categories, invoked_by}
        wf_agg = defaultdict(lambda: {
            "projects": set(), "total_invocations": 0,
            "categories": set(), "invoked_by_projects": set()
        })

        for proj in self.projects:
            pname = proj.results["project_info"]["name"]
            for wf_detail in proj.results["workflow_details"]:
                fname = Path(wf_detail["file"]).name
                wf_agg[fname]["projects"].add(pname)
                for cat in wf_detail["categories"]:
                    wf_agg[fname]["categories"].add(cat)
            for inv in proj.results["invoked_workflows"]:
                wf_name = Path(inv["workflow"]).name
                wf_agg[wf_name]["total_invocations"] += 1
                wf_agg[wf_name]["invoked_by_projects"].add(pname)

        rows = []
        total_projects = len(self.projects)

        for wf_name, data in sorted(wf_agg.items()):
            score, description = self._score_workflow(
                wf_name, data, total_projects
            )
            shared_candidate = score >= SHARED_LIB_THRESHOLD
            rows.append({
                "workflow_file":        wf_name,
                "problem_solved":       description,
                "projects_containing":  len(data["projects"]),
                "projects_list":        ", ".join(sorted(data["projects"])),
                "total_invocations":    data["total_invocations"],
                "activity_categories":  ", ".join(sorted(data["categories"])),
                "reusability_score":    score,
                "score_label":          self._score_label(score),
                "shared_lib_candidate": "✅ YES" if shared_candidate else "—",
                "recommendation":       self._recommend_workflow(score, data, wf_name),
            })

        rows.sort(key=lambda x: -x["reusability_score"])
        return rows

    def _score_workflow(self, wf_name, data, total_projects):
        name_lower = wf_name.lower().replace("_","").replace("-","").replace(".xaml","")
        base_score = 40

        # +20 pts: cross-project presence
        proj_ratio = len(data["projects"]) / max(total_projects, 1)
        base_score += int(proj_ratio * 20)

        # +15 pts: invocation frequency
        inv_score = min(data["total_invocations"] * 5, 15)
        base_score += inv_score

        # +up to 20 pts: workflow name pattern match
        description = f"Workflow file used in {len(data['projects'])} project(s)."
        pattern_boost = 0
        for pattern, desc, boost in WORKFLOW_PATTERNS:
            if re.search(pattern, name_lower):
                description = desc
                pattern_boost = boost
                break
        base_score += pattern_boost

        # category bonus: high-value categories
        for cat in data["categories"]:
            cr = CATEGORY_REUSABILITY.get(cat, (50,""))[0]
            if cr >= 80:
                base_score += 3  # bonus per high-value category

        return min(base_score, 100), description

    def _score_label(self, score):
        if score >= 80: return "🟢 High"
        if score >= 60: return "🟡 Medium"
        if score >= 40: return "🟠 Low-Medium"
        return "🔴 Low"

    def _recommend_workflow(self, score, data, wf_name):
        if score >= 80 and len(data["projects"]) >= 2:
            return "Extract to shared reusable library. Standardise interface (arguments) and publish to Orchestrator."
        if score >= 60 and len(data["projects"]) >= 2:
            return "Good candidate for a shared component. Parameterise and document before sharing."
        if score >= 60:
            return "Consider making this reusable — abstract hardcoded values to arguments."
        if data["total_invocations"] == 0:
            return "Not invoked by other workflows. Review if it can be merged or generalised."
        return "Keep project-specific for now. Monitor usage as more projects are added."

    # ── Activity-category level matrix ────────────────────
    def activity_matrix(self) -> list:
        cat_agg = defaultdict(lambda: {
            "total_usages": 0, "projects_using": set(),
            "unique_activity_types": set()
        })
        for proj in self.projects:
            pname = proj.results["project_info"]["name"]
            for act in proj.results["activities"]:
                cat = act["category"]
                cat_agg[cat]["total_usages"]         += act["count"]
                cat_agg[cat]["projects_using"].add(pname)
                cat_agg[cat]["unique_activity_types"].add(act["raw_tag"])

        rows = []
        total_projects = len(self.projects)
        for cat, data in sorted(cat_agg.items()):
            base_score, cat_desc = CATEGORY_REUSABILITY.get(cat, (50,"General automation activities."))
            # Adjust score by cross-project spread
            spread = len(data["projects_using"]) / max(total_projects, 1)
            adjusted = min(int(base_score * 0.7 + spread * 30), 100)
            rows.append({
                "category":              cat,
                "description":           ACTIVITY_DESCRIPTIONS.get(cat, cat_desc),
                "unique_activity_types": len(data["unique_activity_types"]),
                "total_usages":          data["total_usages"],
                "projects_using":        len(data["projects_using"]),
                "projects_list":         ", ".join(sorted(data["projects_using"])),
                "reusability_score":     adjusted,
                "score_label":           self._score_label(adjusted),
                "shared_lib_candidate":  "✅ YES" if adjusted >= SHARED_LIB_THRESHOLD else "—",
                "recommendation":        self._recommend_category(cat, adjusted, data, total_projects),
            })
        rows.sort(key=lambda x: -x["reusability_score"])
        return rows

    def _recommend_category(self, cat, score, data, total_projects):
        spread = len(data["projects_using"])
        if score >= 80 and spread >= 2:
            return f"Create a '{cat} Utilities' shared library with standardised wrappers."
        if score >= 70:
            return f"Document and template common {cat} patterns for team reuse."
        if cat == "UI Automation" and score < 60:
            return "Selector-dependent — reuse within same application family only."
        if cat == "Control Flow":
            return "Extract TryCatch and exception handling into a standard RE Framework template."
        return "Monitor growth; extract shared patterns when 2+ projects share identical logic."

    # ── Shared Library Recommendations ────────────────────
    def shared_library_recommendations(self, wf_matrix, act_matrix) -> list:
        recs = []

        # From high-scoring workflows used in multiple projects
        seen = set()
        for row in wf_matrix:
            if row["reusability_score"] >= SHARED_LIB_THRESHOLD and row["projects_containing"] >= 2:
                key = row["workflow_file"]
                if key not in seen:
                    seen.add(key)
                    recs.append({
                        "type":           "Workflow",
                        "component":      row["workflow_file"],
                        "score":          row["reusability_score"],
                        "used_in":        row["projects_list"],
                        "problem_solved": row["problem_solved"],
                        "action":         "Extract to shared .xaml, define standard In/Out arguments, publish as NuGet package.",
                    })

        # From high-scoring categories
        for row in act_matrix:
            if row["reusability_score"] >= 80 and row["projects_using"] >= 2:
                recs.append({
                    "type":           "Activity Category",
                    "component":      f"{row['category']} Utilities",
                    "score":          row["reusability_score"],
                    "used_in":        row["projects_list"],
                    "problem_solved": row["description"],
                    "action":         row["recommendation"],
                })

        recs.sort(key=lambda x: -x["score"])
        return recs


# ══════════════════════════════════════════════════════════
#  Repo Discovery
# ══════════════════════════════════════════════════════════
def discover_projects(root_path: Path) -> list:
    """Walk directory tree and find all UiPath project roots (contains project.json)."""
    found = []
    # Check if root itself is a project
    if (root_path / "project.json").exists():
        found.append(root_path)
        return found
    for pjson in sorted(root_path.rglob("project.json")):
        found.append(pjson.parent)
    return found


# ══════════════════════════════════════════════════════════
#  Console Reporting
# ══════════════════════════════════════════════════════════
def print_project_report(analyzer: UiPathProjectAnalyzer):
    pi  = analyzer.results["project_info"]
    sm  = analyzer.results["summary"]
    acts= analyzer.results["activities"]
    apps= sorted(analyzer.results["applications"])
    web_targets = analyzer.results["web_targets"]
    win_targets = analyzer.results["window_targets"]

    print(f"\n{C.CYAN}{'═'*65}{C.RESET}")
    print(f"{C.BOLD}{C.WHITE}  {pi['name']}  —  Project Report{C.RESET}")
    print(f"{C.CYAN}{'═'*65}{C.RESET}")
    print(f"  {C.BOLD}Description:{C.RESET}  {pi['description']}")
    print(f"  {C.BOLD}Version:{C.RESET}      {pi['project_version']}   {C.GREY}|{C.RESET}  {C.BOLD}Main:{C.RESET} {pi['main_workflow']}")
    print(f"\n  XAML files: {sm['total_xaml_files']}  |  Unique activities: {sm['total_activities_unique']}  |  "
          f"Total usages: {sm['total_activity_usages']}  |  Apps detected: {sm['applications_detected']}")

    current_cat = None
    for act in acts:
        if act["category"] != current_cat:
            current_cat = act["category"]
            print(f"\n  {C.MAGENTA}{C.BOLD}▶ {current_cat}{C.RESET}")
            print(f"  {'Activity':<38} {'Count':>5}")
            print(f"  {'─'*38} {'─'*5}")
        name = act["friendly_name"] or act["raw_tag"]
        print(f"  {name:<38} {act['count']:>5}")

    print(f"\n  {C.BOLD}Applications:{C.RESET}")
    for app in apps:
        icon = "🌐" if any(k in app.lower() for k in ["browser","chrome","firefox","edge","web"]) else "🖥️"
        print(f"    {icon}  {app}")

    # ── Web Targets ──────────────────────────────────────────
    if web_targets:
        print(f"\n  {C.BOLD}{C.CYAN}🌐 Web Browser Targets ({len(web_targets)} unique):{C.RESET}")
        print(f"  {'URL':<50} {'Page Title':<35} {'Browser':<18} {'Activity'}")
        print(f"  {'─'*50} {'─'*35} {'─'*18} {'─'*25}")
        for t in web_targets:
            url   = (t["url"][:48]+"..") if len(t["url"])>50 else t["url"]
            title = (t["page_title"][:33]+"..") if len(t["page_title"])>35 else t["page_title"]
            brow  = t["browser"] or "—"
            act   = t["activity"][:24] if t["activity"] else "—"
            print(f"  {C.BLUE}{url:<50}{C.RESET} {title:<35} {brow:<18} {C.GREY}{act}{C.RESET}")

    # ── Window Targets ───────────────────────────────────────
    if win_targets:
        print(f"\n  {C.BOLD}{C.CYAN}🖥️  Windows Desktop Targets ({len(win_targets)} unique):{C.RESET}")
        print(f"  {'Process (EXE)':<28} {'Window Title':<40} {'Activity'}")
        print(f"  {'─'*28} {'─'*40} {'─'*25}")
        for t in win_targets:
            exe   = t["app_exe"] or "—"
            title = (t["window_title"][:38]+"..") if len(t["window_title"])>40 else t["window_title"]
            act   = t["activity"][:24] if t["activity"] else "—"
            print(f"  {C.YELLOW}{exe:<28}{C.RESET} {title:<40} {C.GREY}{act}{C.RESET}")

def print_reusability_report(wf_matrix, act_matrix, shared_libs, total_projects):
    print(f"\n{C.CYAN}{'═'*70}{C.RESET}")
    print(f"{C.BOLD}{C.WHITE}  REUSABILITY MATRIX  ({total_projects} projects scanned){C.RESET}")
    print(f"{C.CYAN}{'═'*70}{C.RESET}")

    print(f"\n{C.BOLD}{C.YELLOW}  ▶ WORKFLOW-LEVEL REUSABILITY{C.RESET}")
    print(f"  {'Workflow':<28} {'Score':>5}  {'Label':<14} {'Projects':>8}  {'Invocations':>11}  {'Shared Lib?'}")
    print(f"  {'─'*28} {'─'*5}  {'─'*14} {'─'*8}  {'─'*11}  {'─'*11}")
    for r in wf_matrix:
        print(f"  {r['workflow_file']:<28} {r['reusability_score']:>5}  {r['score_label']:<14} "
              f"{r['projects_containing']:>8}  {r['total_invocations']:>11}  {r['shared_lib_candidate']}")
        print(f"    {C.GREY}↳ {r['problem_solved']}{C.RESET}")
        print(f"    {C.GREY}  💡 {r['recommendation']}{C.RESET}")

    print(f"\n{C.BOLD}{C.YELLOW}  ▶ ACTIVITY CATEGORY REUSABILITY{C.RESET}")
    print(f"  {'Category':<24} {'Score':>5}  {'Label':<14} {'Projects':>8}  {'Usages':>7}  {'Shared Lib?'}")
    print(f"  {'─'*24} {'─'*5}  {'─'*14} {'─'*8}  {'─'*7}  {'─'*11}")
    for r in act_matrix:
        print(f"  {r['category']:<24} {r['reusability_score']:>5}  {r['score_label']:<14} "
              f"{r['projects_using']:>8}  {r['total_usages']:>7}  {r['shared_lib_candidate']}")
        print(f"    {C.GREY}↳ {r['description']}{C.RESET}")

    if shared_libs:
        print(f"\n{C.BOLD}{C.GREEN}  ▶ SHARED LIBRARY CANDIDATES  ({len(shared_libs)} identified){C.RESET}")
        for i, lib in enumerate(shared_libs, 1):
            print(f"\n  {C.BOLD}{i}. [{lib['type']}] {lib['component']}  (score: {lib['score']}){C.RESET}")
            print(f"     Used in:  {lib['used_in']}")
            print(f"     Solves:   {lib['problem_solved']}")
            print(f"     Action:   {C.CYAN}{lib['action']}{C.RESET}")

    print(f"\n{C.CYAN}{'═'*70}{C.RESET}\n")


# ══════════════════════════════════════════════════════════
#  Excel Export  (full multi-sheet workbook)
# ══════════════════════════════════════════════════════════
def export_excel(projects, wf_matrix, act_matrix, shared_libs, output_dir: Path) -> Path:
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    filename = output_dir / f"{label}_UiPath_Report_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Styles ──
    HDR   = PatternFill("solid", fgColor="1F3864")
    CAT   = PatternFill("solid", fgColor="2E75B6")
    ALT   = PatternFill("solid", fgColor="EBF3FB")
    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
    AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")
    RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
    HDR_F = Font(bold=True, color="FFFFFF", size=11)
    CAT_F = Font(bold=True, color="FFFFFF", size=10)
    TITLE_F = Font(bold=True, size=14, color="1F3864")
    thin  = Side(style="thin", color="BDD7EE")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)

    def hdr(ws, row, cols, fill=HDR, font=HDR_F):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill=fill; cell.font=font
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=bdr
        ws.row_dimensions[row].height=22

    def autowidth(ws, extra=4):
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+extra, 55)

    def score_fill(score):
        if score >= 80: return GREEN_FILL
        if score >= 60: return AMBER_FILL
        return RED_FILL

    # ─────────────────────────────────────────────────
    # Sheet 1 — Repo Summary
    # ─────────────────────────────────────────────────
    ws = wb.active; ws.title = "📊 Repo Summary"
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:D1")
    t=ws["A1"]; t.value="UiPath RPA Repository Analysis"
    t.font=TITLE_F; t.alignment=Alignment(horizontal="center")
    ws.append([])
    ws.append(["Analyzed On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total Projects", len(projects)])
    ws.append(["Total XAML Files", sum(len(p.xaml_files) for p in projects)])
    ws.append(["Total Unique Activity Types",
               len(set(a["raw_tag"] for p in projects for a in p.results["activities"]))])
    ws.append(["Total Activity Usages",
               sum(a["count"] for p in projects for a in p.results["activities"])])
    ws.append([])
    ws.append(["#","Project","Description","Version","XAML Files","Activities","Applications"])
    hdr(ws, ws.max_row, ["#","Project","Description","Version","XAML Files","Activities","Applications"])
    for i, proj in enumerate(projects, 1):
        pi=proj.results["project_info"]; sm=proj.results["summary"]
        ws.append([i, pi["name"], pi["description"], pi["project_version"],
                   sm["total_xaml_files"], sm["total_activity_usages"],
                   sm["applications_detected"]])
        r=ws.max_row
        fill=ALT if i%2==0 else PatternFill()
        for c in range(1,8):
            cell=ws.cell(r,c)
            if fill.patternType: cell.fill=fill
            cell.border=bdr
    autowidth(ws)

    # ─────────────────────────────────────────────────
    # Sheet 2 — Workflow Reusability Matrix
    # ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("🔁 Workflow Matrix")
    ws2.sheet_view.showGridLines=False
    cols=["Workflow File","Problem Solved","Projects Containing","Projects List",
          "Total Invocations","Activity Categories","Reusability Score","Score Label",
          "Shared Lib Candidate","Recommendation"]
    hdr(ws2,1,cols)
    for i, row in enumerate(wf_matrix):
        ws2.append([
            row["workflow_file"], row["problem_solved"],
            row["projects_containing"], row["projects_list"],
            row["total_invocations"], row["activity_categories"],
            row["reusability_score"], row["score_label"],
            row["shared_lib_candidate"], row["recommendation"]
        ])
        r=ws2.max_row
        sf=score_fill(row["reusability_score"])
        ws2.cell(r,7).fill=sf; ws2.cell(r,8).fill=sf
        for c in range(1,11):
            ws2.cell(r,c).border=bdr
            ws2.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
        ws2.row_dimensions[r].height=40
    autowidth(ws2)

    # ─────────────────────────────────────────────────
    # Sheet 3 — Activity Category Matrix
    # ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("📦 Activity Matrix")
    ws3.sheet_view.showGridLines=False
    cols=["Category","Description","Unique Activity Types","Total Usages",
          "Projects Using","Projects List","Reusability Score","Score Label",
          "Shared Lib Candidate","Recommendation"]
    hdr(ws3,1,cols)
    for i, row in enumerate(act_matrix):
        ws3.append([
            row["category"], row["description"],
            row["unique_activity_types"], row["total_usages"],
            row["projects_using"], row["projects_list"],
            row["reusability_score"], row["score_label"],
            row["shared_lib_candidate"], row["recommendation"]
        ])
        r=ws3.max_row
        sf=score_fill(row["reusability_score"])
        ws3.cell(r,7).fill=sf; ws3.cell(r,8).fill=sf
        for c in range(1,11):
            ws3.cell(r,c).border=bdr
            ws3.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
        ws3.row_dimensions[r].height=40
    autowidth(ws3)

    # ─────────────────────────────────────────────────
    # Sheet 4 — Shared Library Candidates
    # ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("⭐ Shared Library Candidates")
    ws4.sheet_view.showGridLines=False
    if shared_libs:
        hdr(ws4,1,["#","Type","Component","Score","Used In Projects","Problem Solved","Recommended Action"])
        for i, lib in enumerate(shared_libs, 1):
            ws4.append([i, lib["type"], lib["component"], lib["score"],
                        lib["used_in"], lib["problem_solved"], lib["action"]])
            r=ws4.max_row
            ws4.cell(r,4).fill=score_fill(lib["score"])
            for c in range(1,8):
                ws4.cell(r,c).border=bdr
                ws4.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
            ws4.row_dimensions[r].height=45
    else:
        ws4["A1"]="No shared library candidates identified yet. Analyse more projects to find patterns."
    autowidth(ws4)

    # ─────────────────────────────────────────────────
    # Sheet 5 — All Activities (cross-project)
    # ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("📋 All Activities")
    ws5.sheet_view.showGridLines=False
    hdr(ws5,1,["Project","Category","Activity Name","Raw Tag","Usage Count","XAML Files"])
    row_num=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for act in proj.results["activities"]:
            files_str=", ".join(Path(f).name for f in act["files"])
            ws5.append([pname, act["category"],
                        act["friendly_name"] or act["raw_tag"],
                        act["raw_tag"], act["count"], files_str])
            r=ws5.max_row; row_num+=1
            fill=ALT if row_num%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws5.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
    autowidth(ws5)

    # ─────────────────────────────────────────────────
    # Sheet 6 — Applications Map
    # ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("🖥️ Applications Map")
    ws6.sheet_view.showGridLines=False
    hdr(ws6,1,["Application","Type","Projects Using It"])
    app_proj = defaultdict(set)
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for app in proj.results["applications"]:
            app_proj[app].add(pname)
    for i, (app, projs) in enumerate(sorted(app_proj.items()), 1):
        atype = "Web" if any(k in app.lower() for k in ["browser","chrome","firefox","edge"]) \
                else "Windows Desktop" if any(k in app.lower() for k in ["excel","word","outlook","sap"]) \
                else "Platform / Service"
        ws6.append([app, atype, ", ".join(sorted(projs))])
        r=ws6.max_row
        fill=ALT if i%2==0 else PatternFill()
        for c in range(1,4):
            cell=ws6.cell(r,c)
            if fill.patternType: cell.fill=fill
            cell.border=bdr
    autowidth(ws6)

    # ─────────────────────────────────────────────────
    # Sheet 7 — Dependencies cross-ref
    # ─────────────────────────────────────────────────
    ws7 = wb.create_sheet("📦 Dependencies")
    ws7.sheet_view.showGridLines=False
    hdr(ws7,1,["Package","Version","Projects Using"])
    pkg_proj = defaultdict(lambda: defaultdict(set))
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for pkg, ver in proj.results["project_info"]["dependencies"].items():
            pkg_proj[pkg][ver].add(pname)
    for i, (pkg, versions) in enumerate(sorted(pkg_proj.items()), 1):
        for ver, projs in versions.items():
            ws7.append([pkg, ver, ", ".join(sorted(projs))])
            r=ws7.max_row
            fill=ALT if i%2==0 else PatternFill()
            for c in range(1,4):
                cell=ws7.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
    autowidth(ws7)

    # ─────────────────────────────────────────────────
    # Sheet 8 — Web Browser Targets (URLs + Page Titles)
    # ─────────────────────────────────────────────────
    ws8 = wb.create_sheet("🌐 Web Targets")
    ws8.sheet_view.showGridLines=False
    hdr(ws8,1,["Project","URL","Page Title","Browser","Activity","XAML File"])
    row_n=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["web_targets"]:
            ws8.append([pname, t["url"], t["page_title"],
                        t["browser"] or "—", t["activity"],
                        Path(t["xaml_file"]).name])
            r=ws8.max_row; row_n+=1
            fill=ALT if row_n%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws8.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
                cell.alignment=Alignment(vertical="center",wrap_text=True)
            # Hyperlink the URL cell
            url_val=t["url"]
            if url_val and url_val.startswith("http"):
                ws8.cell(r,2).hyperlink=url_val
                ws8.cell(r,2).font=Font(color="0563C1",underline="single")
    if row_n==0:
        ws8["A2"]="No web browser targets found. Ensure XAML files contain OpenBrowser/AttachBrowser/Navigate activities with Selector or Url attributes."
    autowidth(ws8)

    # ─────────────────────────────────────────────────
    # Sheet 9 — Windows Desktop Targets
    # ─────────────────────────────────────────────────
    ws9 = wb.create_sheet("🖥️ Window Targets")
    ws9.sheet_view.showGridLines=False
    hdr(ws9,1,["Project","Process (EXE)","Friendly App Name","Window Title","Activity","XAML File"])
    row_n=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["window_targets"]:
            ws9.append([pname, t["app_exe"], t["app_friendly"],
                        t["window_title"], t["activity"],
                        Path(t["xaml_file"]).name])
            r=ws9.max_row; row_n+=1
            fill=ALT if row_n%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws9.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
                cell.alignment=Alignment(vertical="center",wrap_text=True)
    if row_n==0:
        ws9["A2"]="No Windows desktop targets found. Ensure XAML files contain AttachWindow activities with Selector attributes."
    autowidth(ws9)

    wb.save(filename)
    return filename


# ══════════════════════════════════════════════════════════
#  JSON Export
# ══════════════════════════════════════════════════════════
def export_json(projects, wf_matrix, act_matrix, shared_libs, output_dir: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    filename = output_dir / f"{label}_UiPath_Report_{ts}.json"

    def clean(obj):
        if isinstance(obj, set): return sorted(obj)
        if isinstance(obj, dict): return {k: clean(v) for k,v in obj.items()}
        if isinstance(obj, list): return [clean(i) for i in obj]
        return obj

    data = {
        "analyzed_on": datetime.now().isoformat(),
        "total_projects": len(projects),
        "projects": [clean(p.results) for p in projects],
        "reusability_matrix": {
            "workflow_level": wf_matrix,
            "activity_category_level": act_matrix,
            "shared_library_candidates": shared_libs,
        },
        "web_targets": [
            dict(t, project=p.results["project_info"]["name"])
            for p in projects for t in p.results["web_targets"]
        ],
        "window_targets": [
            dict(t, project=p.results["project_info"]["name"])
            for p in projects for t in p.results["window_targets"]
        ],
    }
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    return filename


# ══════════════════════════════════════════════════════════
#  CSV Export
# ══════════════════════════════════════════════════════════
def export_csv(projects, wf_matrix, act_matrix, output_dir: Path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    paths = []

    # Activities CSV
    rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for act in proj.results["activities"]:
            rows.append({
                "Project": pname,
                "Category": act["category"],
                "Activity Name": act["friendly_name"] or act["raw_tag"],
                "Raw Tag": act["raw_tag"],
                "Usage Count": act["count"],
                "XAML Files": ", ".join(Path(f).name for f in act["files"]),
            })
    p1 = output_dir / f"{label}_Activities_{ts}.csv"
    pd.DataFrame(rows).to_csv(p1, index=False); paths.append(p1)

    # Workflow Matrix CSV
    p2 = output_dir / f"{label}_WorkflowMatrix_{ts}.csv"
    pd.DataFrame(wf_matrix).to_csv(p2, index=False); paths.append(p2)

    # Activity Matrix CSV
    p3 = output_dir / f"{label}_ActivityMatrix_{ts}.csv"
    pd.DataFrame(act_matrix).to_csv(p3, index=False); paths.append(p3)

    # Web Targets CSV
    web_rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["web_targets"]:
            web_rows.append({"Project":pname,"URL":t["url"],"Page Title":t["page_title"],
                             "Browser":t["browser"],"Activity":t["activity"],
                             "XAML File":Path(t["xaml_file"]).name})
    p4 = output_dir / f"{label}_WebTargets_{ts}.csv"
    pd.DataFrame(web_rows).to_csv(p4, index=False); paths.append(p4)

    # Window Targets CSV
    win_rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["window_targets"]:
            win_rows.append({"Project":pname,"Process (EXE)":t["app_exe"],
                             "Friendly App Name":t["app_friendly"],
                             "Window Title":t["window_title"],
                             "Activity":t["activity"],
                             "XAML File":Path(t["xaml_file"]).name})
    p5 = output_dir / f"{label}_WindowTargets_{ts}.csv"
    pd.DataFrame(win_rows).to_csv(p5, index=False); paths.append(p5)

    return paths


# ══════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        description="UiPath RPA Analyzer v2 — single project or full repo + reusability matrix",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single project, all outputs
  python3 uipath_analyzer.py ./MyProject --output all --report-dir ./reports

  # Full repository scan
  python3 uipath_analyzer.py ./MyRepo --repo --output all --report-dir ./reports

  # Console only
  python3 uipath_analyzer.py ./MyRepo --repo
        """
    )
    p.add_argument("path", help="Project folder OR repository root (use --repo for multi-project)")
    p.add_argument("--repo",       action="store_true", help="Scan entire repo for multiple projects")
    p.add_argument("--output",     choices=["console","excel","json","csv","all"], default="console")
    p.add_argument("--report-dir", default=".", help="Output directory for files")
    p.add_argument("--quiet",      action="store_true", help="Suppress console report when exporting")
    return p.parse_args()


def main():
    banner()
    args   = parse_args()
    root   = Path(args.path)
    outdir = Path(args.report_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    # ── Discover projects ──────────────────────────────────
    if args.repo:
        print(f"{C.BOLD}[1/5] Scanning repository: {root}{C.RESET}")
        project_paths = discover_projects(root)
        if not project_paths:
            print(f"{C.RED}[ERROR] No UiPath projects (project.json) found under {root}{C.RESET}")
            sys.exit(1)
        print(f"{C.GREEN}[✓] Found {len(project_paths)} project(s){C.RESET}")
    else:
        project_paths = [root]
        print(f"{C.BOLD}[1/5] Single-project mode: {root}{C.RESET}")

    # ── Analyse each project ───────────────────────────────
    print(f"{C.BOLD}[2/5] Analysing projects...{C.RESET}")
    projects = []
    for pp in project_paths:
        analyzer = UiPathProjectAnalyzer(pp)
        if analyzer.run():
            pname = analyzer.results["project_info"]["name"]
            nxaml = len(analyzer.xaml_files)
            print(f"  {C.GREEN}✓{C.RESET}  {pname:<35} {C.GREY}({nxaml} XAML files){C.RESET}")
            projects.append(analyzer)
        else:
            print(f"  {C.YELLOW}⚠{C.RESET}  Skipped (no project.json): {pp}")

    if not projects:
        print(f"{C.RED}[ERROR] No valid projects found.{C.RESET}"); sys.exit(1)

    # ── Reusability matrix ─────────────────────────────────
    print(f"{C.BOLD}[3/5] Computing reusability matrix...{C.RESET}")
    engine    = ReusabilityEngine(projects)
    wf_matrix = engine.workflow_matrix()
    act_matrix= engine.activity_matrix()
    shared    = engine.shared_library_recommendations(wf_matrix, act_matrix)
    print(f"  {C.GREEN}✓{C.RESET}  {len(wf_matrix)} workflows scored  |  "
          f"{len(act_matrix)} categories scored  |  {len(shared)} shared lib candidates")

    # ── Console output ─────────────────────────────────────
    print(f"{C.BOLD}[4/5] Generating reports...{C.RESET}")
    if args.output in ("console","all") and not args.quiet:
        for proj in projects:
            print_project_report(proj)
        print_reusability_report(wf_matrix, act_matrix, shared, len(projects))

    # ── File exports ───────────────────────────────────────
    print(f"{C.BOLD}[5/5] Exporting files...{C.RESET}")
    if args.output in ("excel","all"):
        path = export_excel(projects, wf_matrix, act_matrix, shared, outdir)
        print(f"  {C.GREEN}[✓] Excel  → {path}{C.RESET}")
    if args.output in ("json","all"):
        path = export_json(projects, wf_matrix, act_matrix, shared, outdir)
        print(f"  {C.GREEN}[✓] JSON   → {path}{C.RESET}")
    if args.output in ("csv","all"):
        paths = export_csv(projects, wf_matrix, act_matrix, outdir)
        for path in paths:
            print(f"  {C.GREEN}[✓] CSV    → {path}{C.RESET}")
    if args.output == "console":
        print(f"  {C.GREY}(no files exported — use --output excel/json/csv/all){C.RESET}")

    print(f"\n{C.CYAN}✅  Analysis complete!{C.RESET}\n")


if __name__ == "__main__":
    main()
