#!/usr/bin/env python3
"""
UiPath RPA Project Analyzer  v2.0
===================================
Scans a single project OR an entire repository of UiPath Studio projects.
Generates a full Activity/Application report PLUS a Reusability Matrix with:
  - Usage frequency across all projects
  - Rule-based description of what each workflow/activity solves
  - Reusability score (0–100) per workflow and per activity category
  - Shared library candidate recommendations

Usage:
  # Single project
  python3 uipath_analyzer.py ./MyProject --output all

  # Entire repo (auto-discovers all projects)
  python3 uipath_analyzer.py ./MyRepo --repo --output all --report-dir ./reports

  # Repo, console only
  python3 uipath_analyzer.py ./MyRepo --repo
"""

import os, sys, json, argparse, re
from pathlib import Path
from datetime import datetime
from xml.etree import ElementTree as ET
from collections import defaultdict
from copy import deepcopy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ══════════════════════════════════════════════════════════
#  ANSI colour helpers
# ══════════════════════════════════════════════════════════
class C:
    RESET="\033[0m"; BOLD="\033[1m"; CYAN="\033[96m"; GREEN="\033[92m"
    YELLOW="\033[93m"; RED="\033[91m"; BLUE="\033[94m"; MAGENTA="\033[95m"
    WHITE="\033[97m"; GREY="\033[90m"

def banner():
    print(f"""{C.CYAN}{C.BOLD}
 ██╗   ██╗██╗██████╗  █████╗ ████████╗██╗  ██╗
 ██║   ██║██║██╔══██╗██╔══██╗╚══██╔══╝██║  ██║
 ██║   ██║██║██████╔╝███████║   ██║   ███████║
 ██║   ██║██║██╔═══╝ ██╔══██║   ██║   ██╔══██║
 ╚██████╔╝██║██║     ██║  ██║   ██║   ██║  ██║
  ╚═════╝ ╚═╝╚═╝     ╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝
{C.RESET}{C.BLUE}  UiPath RPA Analyzer + Reusability Matrix  |  v2.2{C.RESET}
""")

# ══════════════════════════════════════════════════════════
#  Activity Knowledge Base
# ══════════════════════════════════════════════════════════
ACTIVITY_CATALOG = {
    "ExcelApplicationScope":("Excel","Excel Application Scope"),
    "ReadRange":("Excel","Read Range"), "WriteRange":("Excel","Write Range"),
    "WriteCell":("Excel","Write Cell"), "ReadCell":("Excel","Read Cell"),
    "AppendRange":("Excel","Append Range"), "SaveWorkbook":("Excel","Save Workbook"),
    "CloseWorkbook":("Excel","Close Workbook"), "InsertRows":("Excel","Insert Rows"),
    "DeleteRows":("Excel","Delete Rows"), "SortRange":("Excel","Sort Range"),
    "FilterTable":("Excel","Filter Table"), "GetWorkbookSheets":("Excel","Get Workbook Sheets"),
    # ── Classic Web/Browser ──────────────────────────────────────────────────
    "OpenBrowser":("Web Browser","Open Browser"), "CloseBrowser":("Web Browser","Close Browser"),
    "Navigate":("Web Browser","Navigate To"), "NavigateTo":("Web Browser","Navigate To"),
    "UploadFile":("Web Browser","Upload File"), "GetText":("Web Browser","Get Text"),
    "InjectJsScript":("Web Browser","Inject JS Script"), "SelectItem":("Web Browser","Select Item"),
    "WaitForPageLoad":("Web Browser","Wait For Page Load"),
    # ── Modern Web/Browser (Use Application/Browser) ─────────────────────────
    "UseApplicationBrowser":("Web Browser","Use Application/Browser [Modern]"),
    "BrowserScope":         ("Web Browser","Browser Scope [Modern]"),
    # ── Classic UI Automation ────────────────────────────────────────────────
    "Click":("UI Automation","Click"), "TypeInto":("UI Automation","Type Into"),
    "SendHotkey":("UI Automation","Send Hotkey"), "AttachWindow":("UI Automation","Attach Window"),
    "AttachBrowser":("UI Automation","Attach Browser"), "FindElement":("UI Automation","Find Element"),
    "WaitElementVanish":("UI Automation","Wait Element Vanish"),
    "DoubleClick":("UI Automation","Double Click"), "RightClick":("UI Automation","Right Click"),
    "Hover":("UI Automation","Hover"), "SetText":("UI Automation","Set Text"),
    "TakeScreenshot":("UI Automation","Take Screenshot"),
    "GetAttribute":("UI Automation","Get Attribute"),
    # ── Modern Desktop (Use Application) ────────────────────────────────────
    "UseApplication":       ("UI Automation","Use Application [Modern]"),
    "ApplicationCard":      ("UI Automation","Application Card [Modern]"),
    "DesktopScope":         ("UI Automation","Desktop Scope [Modern]"),
    "GetOutlookMailMessages":("Outlook/Email","Get Outlook Mail Messages"),
    "SendOutlookMailMessage":("Outlook/Email","Send Outlook Mail Message"),
    "SaveMailMessage":("Outlook/Email","Save Mail Message"),
    "MoveOutlookMailMessage":("Outlook/Email","Move Outlook Mail Message"),
    "MarkAsRead":("Outlook/Email","Mark Mail As Read"),
    "DeleteMailMessage":("Outlook/Email","Delete Mail Message"),
    "SendSmtpMailMessage":("Outlook/Email","Send SMTP Mail Message"),
    "GetImapMailMessages":("Outlook/Email","Get IMAP Mail Messages"),
    "ReadPDFText":("PDF","Read PDF Text"), "ReadPDFWithOCR":("PDF","Read PDF With OCR"),
    "GetPDFPageCount":("PDF","Get PDF Page Count"),
    "JoinPDFFiles":("PDF","Join PDF Files"), "ExtractPDFPage":("PDF","Extract PDF Page"),
    "PresentValidationStation":("Document Understanding","Validation Station"),
    "IntelligentOCRActivity":("Document Understanding","Digitize Document"),
    "ClassifyDocumentScope":("Document Understanding","Classify Document"),
    "ExtractDocumentData":("Document Understanding","Extract Document Data"),
    "SetTransactionStatus":("Orchestrator","Set Transaction Status"),
    "GetTransactionItem":("Orchestrator","Get Transaction Item"),
    "AddQueueItem":("Orchestrator","Add Queue Item"),
    "BulkAddQueueItems":("Orchestrator","Bulk Add Queue Items"),
    "GetQueueItems":("Orchestrator","Get Queue Items"),
    "GetAssetActivity":("Orchestrator","Get Asset"),
    "GetCredentialActivity":("Orchestrator","Get Credential"),
    "SetAssetActivity":("Orchestrator","Set Asset"),
    "LogMessage":("System","Log Message"), "AssignActivity":("System","Assign"),
    "InvokeWorkflowFile":("System","Invoke Workflow File"),
    "InvokeCode":("System","Invoke Code"), "InvokePowerShell":("System","Invoke PowerShell"),
    "StartProcess":("System","Start Process"), "KillProcess":("System","Kill Process"),
    "ReadTextFile":("System","Read Text File"), "WriteTextFile":("System","Write Text File"),
    "AppendLine":("System","Append Line"), "MoveFile":("System","Move File"),
    "CopyFile":("System","Copy File"), "DeleteFile":("System","Delete File"),
    "CompressZipFiles":("System","Compress Zip Files"),
    "HttpClientActivity":("HTTP","HTTP Request"),
    "Connect":("Database","Connect"), "Disconnect":("Database","Disconnect"),
    "ExecuteQuery":("Database","Execute Query"),
    "ExecuteNonQuery":("Database","Execute Non Query"),
    "InsertDataTable":("Database","Insert DataTable"),
    "ForEach":("Control Flow","For Each"), "ForEachRow":("Control Flow","For Each Row"),
    "While":("Control Flow","While"), "If":("Control Flow","If"),
    "Switch":("Control Flow","Switch"), "TryCatch":("Control Flow","Try Catch"),
    "Throw":("Control Flow","Throw"), "Sequence":("Control Flow","Sequence"),
    "Flowchart":("Control Flow","Flowchart"),
    # ── Modern Activities (UiPath 21.10+) ─────────────────────────
    # Use Application/Browser — replaces AttachWindow + OpenBrowser + AttachBrowser
    "UseApplicationBrowser"      :("UI Automation [Modern]","Use Application/Browser"),
    # Modern click/type/extract — child of UseApplicationBrowser
    "ClickActivity"              :("UI Automation [Modern]","Click"),
    "TypeIntoActivity"           :("UI Automation [Modern]","Type Into"),
    "KeyboardShortcutActivity"   :("UI Automation [Modern]","Keyboard Shortcut"),
    "HoverActivity"              :("UI Automation [Modern]","Hover"),
    "SelectItemActivity"         :("UI Automation [Modern]","Select Item"),
    "CheckActivity"              :("UI Automation [Modern]","Check/Uncheck"),
    "GetTextActivity"            :("UI Automation [Modern]","Get Text"),
    "GetAttributeActivity"       :("UI Automation [Modern]","Get Attribute"),
    "SetTextActivity"            :("UI Automation [Modern]","Set Text"),
    "FindChildrenActivity"       :("UI Automation [Modern]","Find Children"),
    "WaitForActivity"            :("UI Automation [Modern]","Wait For"),
    "VerifyActivity"             :("UI Automation [Modern]","Verify Expression with Operator"),
    "TakeScreenshotActivity"     :("UI Automation [Modern]","Take Screenshot"),
    "HighlightActivity"          :("UI Automation [Modern]","Highlight"),
    "DragAndDropActivity"        :("UI Automation [Modern]","Drag And Drop"),
    "ScrollActivity"             :("UI Automation [Modern]","Scroll"),
    "ExtractTableDataActivity"   :("UI Automation [Modern]","Extract Table Data"),
    "InvokeJavaScriptActivity"   :("UI Automation [Modern]","Invoke JavaScript"),
    # Modern Excel (WorkbookApplication / ExcelProcessScope)
    "ExcelProcessScope"          :("Excel [Modern]","Excel Process Scope"),
    "WorkbookApplication"        :("Excel [Modern]","Use Excel File"),
    "ReadRangeActivity"          :("Excel [Modern]","Read Range"),
    "WriteRangeActivity"         :("Excel [Modern]","Write Range"),
    "AppendRangeActivity"        :("Excel [Modern]","Append Range"),
    "GetTableRangeActivity"      :("Excel [Modern]","Get Table Range"),
    "SaveWorkbookActivity"       :("Excel [Modern]","Save Workbook"),
    "InsertColumnsActivity"      :("Excel [Modern]","Insert Columns"),
    "DeleteColumnsActivity"      :("Excel [Modern]","Delete Columns"),
    "AutoFillActivity"           :("Excel [Modern]","Auto Fill"),
    "CopyPasteRangeActivity"     :("Excel [Modern]","Copy/Paste Range"),
    "ForEachExcelRowActivity"    :("Excel [Modern]","For Each Excel Row"),
    # Modern Mail
    "UseDesktopOutlookApp"       :("Outlook/Email [Modern]","Use Desktop Outlook App"),
    "GetEmailActivity"           :("Outlook/Email [Modern]","Get Email"),
    "SendEmailActivity"          :("Outlook/Email [Modern]","Send Email"),
    "MoveEmailActivity"          :("Outlook/Email [Modern]","Move Email"),
    "ReplyToEmailActivity"       :("Outlook/Email [Modern]","Reply To Email"),
    "DeleteEmailActivity"        :("Outlook/Email [Modern]","Delete Email"),
    "SaveEmailActivity"          :("Outlook/Email [Modern]","Save Email"),
    "MarkEmailAsReadActivity"    :("Outlook/Email [Modern]","Mark Email As Read"),
    # Modern File / Folder
    "CopyFileActivity"           :("System [Modern]","Copy File"),
    "MoveFileActivity"           :("System [Modern]","Move File"),
    "DeleteFileActivity"         :("System [Modern]","Delete File"),
    "CreateFolderActivity"       :("System [Modern]","Create Folder"),
    "PathExistsActivity"         :("System [Modern]","Path Exists"),
    "CompressZipActivity"        :("System [Modern]","Compress Zip"),
    "ExtractZipActivity"         :("System [Modern]","Extract Zip"),
    "ForEachFileActivity"        :("System [Modern]","For Each File in Folder"),
    "ForEachFolderActivity"      :("System [Modern]","For Each Folder"),
    "ReadTextFileActivity"       :("System [Modern]","Read Text File"),
    "WriteTextFileActivity"      :("System [Modern]","Write Text File"),
    "AppendLineActivity"         :("System [Modern]","Append Line"),
    # Modern Orchestrator
    "AddQueueItemActivity"       :("Orchestrator [Modern]","Add Queue Item"),
    "GetQueueItemActivity"       :("Orchestrator [Modern]","Get Queue Item"),
    "SetTransactionStatusActivity":("Orchestrator [Modern]","Set Transaction Status"),
    "BulkAddQueueItemsActivity"  :("Orchestrator [Modern]","Bulk Add Queue Items"),
    "GetAssetValueActivity"      :("Orchestrator [Modern]","Get Asset"),
    "SetAssetValueActivity"      :("Orchestrator [Modern]","Set Asset"),
    "GetCredentialsActivity"     :("Orchestrator [Modern]","Get Credentials"),
    # Modern PDF
    "ReadPdfTextActivity"        :("PDF [Modern]","Read PDF Text"),
}

APP_MAP = {
    "Excel":"Microsoft Excel", "Web Browser":"Web Browser (Chrome/Edge/Firefox)",
    "Outlook/Email":"Microsoft Outlook / Email", "PDF":"PDF Files",
    "Document Understanding":"UiPath Document Understanding / IDP",
    "Orchestrator":"UiPath Orchestrator", "SAP":"SAP ERP",
    "Microsoft Word":"Microsoft Word", "Salesforce":"Salesforce CRM",
    "Database":"Database (SQL/OLEDB)", "HTTP":"HTTP / REST APIs",
    "UI Automation":"Windows Desktop Application",
    # Modern equivalents map to same apps
    "UI Automation [Modern]":"Windows Desktop / Web Application (Modern)",
    "Excel [Modern]":"Microsoft Excel",
    "Outlook/Email [Modern]":"Microsoft Outlook / Email",
    "System [Modern]":"System (File/Folder Operations)",
    "Orchestrator [Modern]":"UiPath Orchestrator",
    "PDF [Modern]":"PDF Files",
}

WINDOWS_APP_PATTERNS = {
    r"sap":"SAP ERP (Windows GUI)", r"outlook":"Microsoft Outlook",
    r"excel":"Microsoft Excel", r"winword":"Microsoft Word",
    r"chrome":"Google Chrome", r"firefox":"Mozilla Firefox",
    r"msedge|edge":"Microsoft Edge", r"acrobat":"Adobe Acrobat",
    r"salesforce":"Salesforce (Browser)", r"servicenow":"ServiceNow",
    r"workday":"Workday",
}

# ══════════════════════════════════════════════════════════
#  Reusability Knowledge Base  (rule-based, fully offline)
# ══════════════════════════════════════════════════════════

# Per-category: base reusability score (0-100) + description of value
CATEGORY_REUSABILITY = {
    "Orchestrator":      (92, "Queue & transaction management; forms the RE Framework backbone reusable across any automation."),
    "Orchestrator [Modern]": (92, "Modern queue & credential activities — same high reusability as classic Orchestrator."),
    "System":            (88, "Core utilities (logging, file ops, workflow invocation) applicable to virtually every RPA project."),
    "System [Modern]":   (88, "Modern file/folder activities — functionally identical reusability to classic System activities."),
    "Outlook/Email":     (85, "Email notification & retrieval patterns that repeat in most business process automations."),
    "Outlook/Email [Modern]": (85, "Modern Outlook activities — same high reusability as classic mail activities."),
    "Excel":             (82, "Spreadsheet read/write operations commonly shared across Finance, HR and Operations workflows."),
    "Excel [Modern]":    (82, "Modern Excel activities — same reusability profile as classic Excel; prefer for new projects."),
    "PDF":               (78, "Document extraction logic reusable wherever PDF invoices, receipts or reports are processed."),
    "PDF [Modern]":      (78, "Modern PDF extraction — same reusability as classic PDF activities."),
    "Document Understanding": (75, "IDP extraction pipeline reusable across invoice, PO, and contract processing projects."),
    "Database":          (80, "Database connectivity & query patterns shareable across all data-intensive automations."),
    "HTTP":              (77, "REST API integration wrappers reusable across any project consuming web services."),
    "Web Browser":       (60, "Browser automation steps — reusable when targeting the same web application."),
    "UI Automation":     (50, "Screen interaction steps — lower reusability as they are selector-dependent per application."),
    "UI Automation [Modern]": (55, "Modern UI automation — slightly higher reusability than classic due to descriptive targeting."),
    "Control Flow":      (70, "Standard error handling & loop patterns (TryCatch, ForEach) reusable as framework templates."),
    "SAP":               (65, "SAP GUI automation steps — reusable within SAP-focused project families."),
    "Microsoft Word":    (72, "Word document generation patterns reusable in reporting workflows."),
    "Salesforce":        (68, "Salesforce CRM interaction patterns reusable across Sales/Support automations."),
}

# Workflow-name pattern → (problem it solves, reusability boost 0-20)
WORKFLOW_PATTERNS = [
    (r"send.*email|email.*notif|notification",  "Sends automated email notifications on process completion or failure.", 20),
    (r"error.*handl|exception.*handl|handle.*error", "Centralised exception handling and alerting logic.", 18),
    (r"login|authenticate|credential",          "Handles application login and credential retrieval.", 17),
    (r"read.*excel|excel.*read|load.*data",     "Reads and loads structured data from Excel spreadsheets.", 15),
    (r"write.*excel|excel.*write|save.*data",   "Writes or updates processed data back to Excel.", 14),
    (r"extract.*pdf|pdf.*extract|ocr",          "Extracts text or structured data from PDF documents using OCR.", 16),
    (r"validate|validation",                    "Validates extracted or input data against business rules.", 15),
    (r"queue|transaction",                      "Manages Orchestrator queue items and transaction lifecycle.", 18),
    (r"log|logging",                            "Provides centralised logging and audit trail.", 16),
    (r"report|summary",                         "Generates and distributes process summary reports.", 14),
    (r"sap.*post|post.*sap|sap.*entry",         "Posts data entries into SAP ERP modules.", 13),
    (r"ad.*account|active.*directory",          "Creates or manages Active Directory user accounts.", 14),
    (r"db.*connect|database.*connect|connect",  "Establishes and manages database connections.", 15),
    (r"download|fetch.*file|get.*file",         "Downloads or retrieves files from remote sources.", 13),
    (r"archive|move.*file|cleanup",             "Archives processed files and performs cleanup operations.", 12),
    (r"config|setting|parameter",               "Loads runtime configuration and parameters.", 17),
    (r"main|orchestrat",                        "Orchestrates the end-to-end process flow calling sub-workflows.", 10),
]

# Activity tag patterns → problem description
ACTIVITY_DESCRIPTIONS = {
    "Orchestrator":  "Manages process queues, credentials and transaction state via UiPath Orchestrator.",
    "System":        "Handles file operations, logging, and invocation of modular sub-workflows.",
    "Outlook/Email": "Automates email retrieval, dispatch and inbox management via Microsoft Outlook.",
    "Excel":         "Reads, writes and manipulates data in Microsoft Excel workbooks.",
    "PDF":           "Extracts and processes text and structured data from PDF documents.",
    "Document Understanding": "Applies intelligent document processing (IDP) to classify and extract unstructured data.",
    "Database":      "Connects to and queries relational databases for data retrieval or updates.",
    "HTTP":          "Integrates with external REST/SOAP APIs via HTTP requests.",
    "Web Browser":   "Automates interactions with web-based applications in the browser.",
    "UI Automation": "Drives desktop or web UI elements (clicks, typing, reading values).",
    "Control Flow":  "Provides structured error handling, loops and conditional branching.",
    "SAP":           "Automates SAP ERP GUI transactions and data entry.",
    "Microsoft Word":"Creates or manipulates Microsoft Word documents programmatically.",
    "Salesforce":    "Automates Salesforce CRM record management and queries.",
}

SHARED_LIB_THRESHOLD = 60   # score >= this → candidate for shared library


# ══════════════════════════════════════════════════════════
#  Single-Project Analyzer
# ══════════════════════════════════════════════════════════
class UiPathProjectAnalyzer:
    def __init__(self, project_path: Path):
        self.project_path = Path(project_path).resolve()
        self.project_json = {}
        self.xaml_files   = []
        self.results = {
            "project_info": {}, "activities": [],
            "applications": set(), "web_urls": set(),
            "window_selectors": set(), "invoked_workflows": [],
            "summary": {}, "unknown_activities": [],
            "workflow_details": [],          # per-XAML file breakdown
            # ── NEW: deep target extraction ──────────────────────
            "web_targets":     [],   # {url, page_title, browser, activity, xaml_file}
            "window_targets":  [],   # {app_exe, window_title, activity, xaml_file}
        }

    def discover(self):
        pj = self.project_path / "project.json"
        if not pj.exists():
            return False
        with open(pj, "r", encoding="utf-8") as f:
            self.project_json = json.load(f)
        self.xaml_files = sorted(self.project_path.rglob("*.xaml"))
        return True

    def parse_all_xaml(self):
        activity_map = defaultdict(lambda: {
            "count":0,"category":"","friendly_name":"","raw_tag":"","files":set()
        })
        for xf in self.xaml_files:
            rel = xf.relative_to(self.project_path)
            self._parse_xaml(xf, rel, activity_map)
        for raw_tag, info in activity_map.items():
            info["files"] = sorted(info["files"])
            info["raw_tag"] = raw_tag
            self.results["activities"].append(dict(info))
        self.results["activities"].sort(key=lambda x:(x["category"],-x["count"]))

    def _parse_xaml(self, xaml_path, rel_path, activity_map):
        try:
            tree = ET.parse(xaml_path)
            root = tree.getroot()
        except ET.ParseError:
            return
        # Per-workflow detail
        wf_detail = {
            "file": str(rel_path), "display_name": root.get("DisplayName",""),
            "activities": defaultdict(int), "categories": set(),
            "invokes": []
        }
        self._walk_element(root, str(rel_path), activity_map, wf_detail)
        wf_detail["activities"] = dict(wf_detail["activities"])
        wf_detail["categories"] = sorted(wf_detail["categories"])
        self.results["workflow_details"].append(wf_detail)

    def _local_name(self, tag):
        return tag.split("}")[-1] if "}" in tag else tag

    def _walk_element(self, elem, filename, activity_map, wf_detail):
        local = self._local_name(elem.tag)
        if local not in ("Activity","Members","x:Members"):
            cat, friendly = self._classify_activity(local, elem)
            if cat:
                activity_map[local]["count"]        += 1
                activity_map[local]["category"]      = cat
                activity_map[local]["friendly_name"] = friendly
                activity_map[local]["files"].add(filename)
                wf_detail["activities"][local] = wf_detail["activities"].get(local,0)+1
                wf_detail["categories"].add(cat)
                app = APP_MAP.get(cat)
                if app:
                    self.results["applications"].add(app)

                display_name = elem.get("DisplayName", local)

                # ── URL extraction (direct attributes) — Classic + Modern ──
                for url_attr in ("Url","url","URL","UrlString","BrowserURL","NavigateUrl"):
                    url_val = elem.get(url_attr, "")
                    if url_val and url_val.startswith("http"):
                        self.results["web_urls"].add(url_val)
                        self._record_web_target(
                            url=url_val, page_title="",
                            browser=self._guess_browser(elem),
                            activity=display_name, xaml_file=filename,
                            activity_type=local
                        )

                # ── Classic Selector attribute ───────────────────
                sel_raw = elem.get("Selector") or elem.get("selector") or ""
                if sel_raw:
                    self._deep_parse_selector(sel_raw, display_name, filename, local)

                # ── Modern: UseApplication (desktop scope) ────────
                if local in ("UseApplication","ApplicationCard","DesktopScope"):
                    self._extract_modern_use_application(elem, display_name, filename)

                # ── Modern: UseBrowser / UseApplicationBrowser ────
                elif local in ("UseBrowser","UseApplicationBrowser","BrowserScope","UseApp"):
                    self._extract_modern_use_browser(elem, display_name, filename)
                    self._extract_modern_use_app_browser(elem, display_name, filename)

                # ── Modern: NApplicationCard (real-world v23+ scope) ─
                elif local == "NApplicationCard":
                    self._extract_napplication_card(elem, display_name, filename)

                # ── Modern: NavigateTo / BrowserNavigateTo ────────
                elif local in ("NavigateTo","BrowserNavigateTo","NavigateToActivity",
                               "NNavigateBrowser"):
                    url_val = elem.get("Url","") or elem.get("URL","")
                    if not url_val:
                        url_val = self._extract_inargument_text(elem)
                    if url_val and url_val.startswith("http"):
                        self.results["web_urls"].add(url_val)
                        self._record_web_target(
                            url=url_val, page_title="",
                            browser=self._guess_browser(elem),
                            activity=display_name, xaml_file=filename,
                            activity_type=local
                        )

                # ── Modern Target child element ───────────────────
                self._extract_modern_target(elem, display_name, filename, local)

                # ── TargetAnchorable: extract BrowserURL + ScopeSelectorArgument ──
                # Handles NGetText, NClick, NTypeInto, NCheckState etc.
                self._extract_target_anchorable(elem, display_name, filename, local)

                # ── InvokeWorkflowFile ────────────────────────────
                if local in ("InvokeWorkflowFile","InvokeWorkflowFileActivity"):
                    wf = (elem.get("InvokedWorkflowFileName") or
                          elem.get("WorkflowFileName") or "")
                    if wf:
                        self.results["invoked_workflows"].append(
                            {"workflow":wf,"called_from":filename})
                        wf_detail["invokes"].append(wf)

        for child in elem:
            self._walk_element(child, filename, activity_map, wf_detail)

    # ── Modern: NApplicationCard (UiPath v23+ browser scope) ─────
    def _extract_napplication_card(self, elem, display_name: str, filename: str):
        """
        NApplicationCard is the real modern scope activity used in UiPath 23.x+.
        Target info lives in a child <uix:NApplicationCard.TargetApp> node:
          <uix:TargetApp BrowserType="Edge"
                         Selector="&lt;html app='msedge.exe' title='...' /&gt;"
                         Title="Organization Details | UiPath Licensing"
                         Url="[Current_URL]" />
        The Url may be a VB expression like [Current_URL] — we skip those.
        ScopeSelectorArgument on TargetAnchorable children carries the html/wnd selector.
        """
        for child in elem:
            child_local = self._local_name(child.tag)
            if child_local in ("NApplicationCard.TargetApp", "TargetApp.TargetApp"):
                for grandchild in child:
                    gc_local = self._local_name(grandchild.tag)
                    if gc_local == "TargetApp":
                        title        = grandchild.get("Title","")
                        browser_type = grandchild.get("BrowserType","")
                        url          = grandchild.get("Url","")
                        selector     = grandchild.get("Selector","")
                        app_path     = grandchild.get("ApplicationPath","")

                        # Skip VB/C# expressions
                        if url and url.startswith("["):
                            url = ""

                        if browser_type or (selector and "html" in selector.lower()):
                            # Web target
                            if not url and selector:
                                # Parse URL from selector
                                sel_decoded = selector.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&").replace("&quot;",'"')
                                url_match = re.search(r"url='([^']+)'", sel_decoded, re.I)
                                if url_match:
                                    url = url_match.group(1)
                            if url and url.startswith("http"):
                                self.results["web_urls"].add(url)
                            if url or title or browser_type:
                                self._record_web_target(
                                    url=url, page_title=title,
                                    browser=self._map_browser_type(browser_type),
                                    activity=display_name, xaml_file=filename,
                                    activity_type="NApplicationCard"
                                )
                            if selector:
                                self._deep_parse_selector(selector, display_name, filename, "NApplicationCard")
                        elif app_path or (selector and "wnd" in selector.lower()):
                            # Desktop target
                            app_exe = Path(app_path).name if app_path else ""
                            self._record_window_target(app_exe, title, display_name, filename, "NApplicationCard")
                            if selector:
                                self._deep_parse_selector(selector, display_name, filename, "NApplicationCard")

    # ── TargetAnchorable: BrowserURL + ScopeSelectorArgument extraction ───
    def _extract_target_anchorable(self, elem, display_name: str, filename: str, activity_type: str):
        """
        Modern N-prefixed activities (NGetText, NClick, NTypeInto, NCheckState etc.)
        store target info in nested <uix:TargetAnchorable> elements with:
          BrowserURL="https://license.uipath.com/license/..."
          ScopeSelectorArgument="&lt;html app='msedge.exe' title='...' /&gt;"

        We recursively look for TargetAnchorable nodes within .Target children
        and extract:
          - BrowserURL → direct URL of the page being automated
          - ScopeSelectorArgument → html/wnd selector for the scope context
        """
        self._scan_for_target_anchorable(elem, display_name, filename, activity_type)

    def _scan_for_target_anchorable(self, elem, display_name, filename, activity_type, depth=0):
        """Recursively scan for TargetAnchorable nodes (max depth 6)."""
        if depth > 6:
            return
        local = self._local_name(elem.tag)

        if local == "TargetAnchorable":
            browser_url = elem.get("BrowserURL","")
            scope_sel   = elem.get("ScopeSelectorArgument","")
            title_attr  = elem.get("Title","") or elem.get("WindowTitle","")

            if browser_url and browser_url.startswith("http"):
                self.results["web_urls"].add(browser_url)
                # Extract page title from ScopeSelectorArgument if present
                page_title = title_attr
                browser    = ""
                if scope_sel:
                    sel_dec = scope_sel.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&").replace("&quot;",'"')
                    title_m  = re.search(r"title='([^']+)'", sel_dec, re.I)
                    app_m    = re.search(r"app='([^']+)'", sel_dec, re.I)
                    if title_m and not page_title: page_title = title_m.group(1)
                    if app_m: browser = app_m.group(1)
                self._record_web_target(
                    url=browser_url, page_title=page_title,
                    browser=self._map_browser_type(browser) if "edge" in browser.lower() or "chrome" in browser.lower() else browser,
                    activity=display_name, xaml_file=filename,
                    activity_type=activity_type
                )

            elif scope_sel:
                # No BrowserURL — parse scope selector for context
                sel_dec = scope_sel.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&").replace("&quot;",'"')
                if "<html" in sel_dec or "<web" in sel_dec:
                    title_m = re.search(r"title='([^']+)'", sel_dec, re.I)
                    app_m   = re.search(r"app='([^']+)'", sel_dec, re.I)
                    url_m   = re.search(r"url='([^']+)'", sel_dec, re.I)
                    self._record_web_target(
                        url=url_m.group(1) if url_m else "",
                        page_title=title_m.group(1) if title_m else title_attr,
                        browser=app_m.group(1) if app_m else "",
                        activity=display_name, xaml_file=filename,
                        activity_type=activity_type
                    )
                elif "<wnd" in sel_dec:
                    app_m   = re.search(r"app='([^']+)'", sel_dec, re.I)
                    title_m = re.search(r"title='([^']+)'", sel_dec, re.I)
                    self._record_window_target(
                        app_exe=app_m.group(1) if app_m else "",
                        window_title=title_m.group(1) if title_m else title_attr,
                        activity=display_name, xaml_file=filename,
                        activity_type=activity_type
                    )

        for child in elem:
            self._scan_for_target_anchorable(child, display_name, filename, activity_type, depth+1)

    # ── Modern: UseApplication (desktop) ──────────────────────────
    def _extract_modern_use_application(self, elem, display_name: str, filename: str):
        """
        Modern UseApplication (desktop scope). Carries target info as:
          ApplicationPath="C:\\...\\saplogon.exe"  (full path — strip to basename)
          Title="SAP Easy Access"
        OR inside child <ui:UseApplication.Target><ui:TargetAnchor>
        """
        title    = elem.get("Title","") or elem.get("WindowTitle","")
        app_path = elem.get("ApplicationPath","") or elem.get("AppPath","")
        # Always strip to basename — avoids full-path duplicate vs selector-derived basename
        app_exe  = Path(app_path).name if app_path else ""

        if app_exe or title:
            self._record_window_target(app_exe, title, display_name, filename,
                                       activity_type="UseApplication")

        # Also check child .Target nodes
        self._scan_target_children(elem, display_name, filename, "UseApplication")

    # ── Modern: UseBrowser / UseApplicationBrowser (web scope) ────
    def _extract_modern_use_browser(self, elem, display_name: str, filename: str):
        """
        Modern UseBrowser / UseApplicationBrowser. Carries:
          Url="https://..."            — direct attribute
          BrowserType="Chrome|Edge|Firefox"
          Title="Page Title"
        OR Url inside <ui:UseBrowser.Url><InArgument x:TypeArguments="x:String">https://...
        """
        title        = elem.get("Title","") or elem.get("WindowTitle","")
        browser_type = elem.get("BrowserType","") or elem.get("Browser","")
        url_val      = (elem.get("Url","") or elem.get("URL","") or
                        elem.get("BrowserURL",""))

        # Try InArgument child text if URL not as attribute
        if not url_val:
            url_val = self._extract_inargument_url_child(elem)

        if url_val and url_val.startswith("http"):
            self.results["web_urls"].add(url_val)
            self._record_web_target(
                url=url_val, page_title=title,
                browser=self._map_browser_type(browser_type),
                activity=display_name, xaml_file=filename,
                activity_type="UseBrowser"
            )
        elif title:
            self._record_web_target(
                url="", page_title=title,
                browser=self._map_browser_type(browser_type),
                activity=display_name, xaml_file=filename,
                activity_type="UseBrowser"
            )

        # Also scan .Target children
        self._scan_target_children(elem, display_name, filename, "UseBrowser")

    def _map_browser_type(self, bt: str) -> str:
        """Map BrowserType string → friendly browser name."""
        mapping = {
            "chrome":"Google Chrome","firefox":"Mozilla Firefox",
            "edge":"Microsoft Edge","ie":"Internet Explorer",
            "internetexplorer":"Internet Explorer",
        }
        return mapping.get(bt.lower(), bt) if bt else ""

    def _extract_inargument_url_child(self, elem) -> str:
        """
        Extract URL from pattern:
          <ui:UseBrowser.Url>
            <InArgument x:TypeArguments="x:String">https://...</InArgument>
          </ui:UseBrowser.Url>
        """
        for child in elem:
            child_local = self._local_name(child.tag)
            if child_local.endswith(".Url") or child_local.endswith(".URL"):
                return self._extract_inargument_text(child)
        return ""

    def _extract_inargument_text(self, elem) -> str:
        """
        Extract plain-text content from InArgument / x:String child nodes.
        Handles: <InArgument ...>value</InArgument>
                 <x:String>value</x:String>
        """
        for child in elem:
            child_local = self._local_name(child.tag)
            if child_local in ("InArgument","String","x:String"):
                if child.text and child.text.strip():
                    return child.text.strip()
        # Also check elem text directly
        if elem.text and elem.text.strip():
            return elem.text.strip()
        return ""

    def _scan_target_children(self, elem, display_name: str, filename: str, activity_type: str):
        """
        Scan child elements of Modern scope activities for embedded Target/Selector nodes.
        Handles:
          <ui:UseApplication.Target>
            <ui:TargetAnchor>
              <ui:TargetAnchor.Selector>
                <x:String>&lt;wnd app='...' title='...' /&gt;</x:String>
              </ui:TargetAnchor.Selector>
            </ui:TargetAnchor>
          </ui:UseApplication.Target>
        """
        for child in elem:
            child_local = self._local_name(child.tag)

            # .Target wrapper
            if child_local.endswith(".Target") or child_local == "Target":
                for grandchild in child:
                    gc_local = self._local_name(grandchild.tag)
                    if gc_local in ("Target","TargetAnchor"):
                        # Direct selector attribute on Target
                        sel = grandchild.get("Selector","")
                        if sel:
                            self._deep_parse_selector(sel, display_name, filename, activity_type)
                        # Selector as child x:String text
                        sel_text = self._extract_selector_from_children(grandchild)
                        if sel_text:
                            self._deep_parse_selector(sel_text, display_name, filename, activity_type)
                        # Title / ApplicationPath on Target node itself
                        title = grandchild.get("Title","") or grandchild.get("WindowTitle","")
                        app   = grandchild.get("ApplicationPath","") or grandchild.get("AppPath","")
                        url   = grandchild.get("Url","") or grandchild.get("URL","")
                        if title and not sel and not sel_text:
                            if url and url.startswith("http"):
                                self._record_web_target(url, title, "", display_name, filename, activity_type)
                            else:
                                self._record_window_target(Path(app).name if app else "", title,
                                                           display_name, filename, activity_type)

    def _extract_selector_from_children(self, elem) -> str:
        """
        Extract selector string from child patterns like:
          <ui:TargetAnchor.Selector><x:String>&lt;wnd .../&gt;</x:String></ui:TargetAnchor.Selector>
        """
        for child in elem:
            child_local = self._local_name(child.tag)
            if child_local.endswith(".Selector") or child_local == "Selector":
                # The value is typically in a nested <x:String> text node
                text = self._extract_inargument_text(child)
                if text:
                    return text
                if child.text and child.text.strip():
                    return child.text.strip()
        return ""

    # ── Modern: UseApplicationBrowser direct attributes ───────────
    def _extract_modern_use_app_browser(self, elem, display_name: str, filename: str):
        """
        UseApplicationBrowser carries target info as direct XML attributes:
          Title="SAP Easy Access"
          ApplicationPath="C:\\Program Files\\SAP\\saplogon.exe"
          Url="https://app.example.com"
          BrowserType="Chrome|Edge|Firefox"
          IsWeb="True" / IsDesktop="True"
        Decision rule: BrowserType → always web. ApplicationPath → always desktop.
        URL → web. IsWeb → web. Else classify by title keywords.
        """
        title        = elem.get("Title","") or elem.get("WindowTitle","")
        app_path     = elem.get("ApplicationPath","") or elem.get("AppPath","")
        url          = elem.get("Url","") or elem.get("URL","") or elem.get("BrowserURL","")
        browser_type = elem.get("BrowserType","") or elem.get("Browser","")
        is_web_attr  = elem.get("IsWeb","").lower() == "true"
        is_desk_attr = elem.get("IsDesktop","").lower() == "true"
        app_exe      = Path(app_path).name if app_path else ""

        if browser_type:
            # BrowserType present → always web regardless of other attrs
            if not url:
                url = self._extract_inargument_url_child(elem)
            if url and url.startswith("http"):
                self.results["web_urls"].add(url)
            self._record_web_target(
                url=url, page_title=title,
                browser=self._map_browser_type(browser_type),
                activity=display_name, xaml_file=filename,
                activity_type="UseApplicationBrowser"
            )
        elif app_exe or is_desk_attr:
            # ApplicationPath present → always desktop
            self._record_window_target(
                app_exe=app_exe, window_title=title,
                activity=display_name, xaml_file=filename,
                activity_type="UseApplicationBrowser"
            )
        elif url and url.startswith("http"):
            # URL but no BrowserType → treat as web
            self.results["web_urls"].add(url)
            self._record_web_target(
                url=url, page_title=title, browser="",
                activity=display_name, xaml_file=filename,
                activity_type="UseApplicationBrowser"
            )
        elif is_web_attr and title:
            self._record_web_target(
                url="", page_title=title, browser="",
                activity=display_name, xaml_file=filename,
                activity_type="UseApplicationBrowser"
            )
        elif title:
            # Classify by title keywords only
            browser_keys = ["portal","workday","salesforce","servicenow","sharepoint",
                            "chrome","firefox","edge","browser","web","http"]
            desktop_keys = ["sap","notepad","excel","word","explorer","cmd","erp",
                            "finance","approval","management system"]
            tl = title.lower()
            if any(k in tl for k in browser_keys):
                self._record_web_target("", title, "",
                                        display_name, filename, "UseApplicationBrowser")
            elif any(k in tl for k in desktop_keys):
                self._record_window_target("", title,
                                           display_name, filename, "UseApplicationBrowser")

    # ── Modern: Target child element extraction ────────────────────
    def _extract_modern_target(self, elem, display_name: str, filename: str, local: str):
        """
        Modern activities store target info in child XML elements:
          <ui:ClickActivity.Target>
            <ui:Target Selector="..." Title="..." ApplicationPath="...">
              <ui:Target.Anchor> ... </ui:Target.Anchor>
            </ui:Target>
          </ui:ClickActivity.Target>
        Also handles:
          Target.Selector="..." as a direct attribute (serialised form)
        """
        # Check for Target.Selector / Target.Title as direct attrs (compact serialisation)
        for attr_key in elem.attrib:
            attr_local = attr_key.split("}")[-1] if "}" in attr_key else attr_key
            if attr_local in ("Target.Selector",):
                sel = elem.get(attr_key,"")
                if sel:
                    self._deep_parse_selector(sel, display_name, filename, local)

        # Walk child nodes looking for Target / Target.* containers
        for child in elem:
            child_local = self._local_name(child.tag)

            # <ActivityName.Target> wrapper node
            if child_local.endswith(".Target") or child_local == "Target":
                self._process_target_node(child, display_name, filename, local)

            # <ui:Target Selector="..." Title="..." ApplicationPath="...">
            if child_local == "Target":
                self._process_target_node(child, display_name, filename, local)

    def _process_target_node(self, target_elem, display_name: str, filename: str, local: str):
        """Extract selector/title/app from a <Target> or <ActivityName.Target> node."""
        for node in target_elem:
            node_local = self._local_name(node.tag)
            if node_local == "Target":
                sel   = node.get("Selector","")
                title = node.get("Title","") or node.get("WindowTitle","")
                app   = node.get("ApplicationPath","") or node.get("AppPath","")
                url   = node.get("Url","") or node.get("URL","")

                if sel:
                    self._deep_parse_selector(sel, display_name, filename, local)
                if title and not sel:
                    # Distinguish web vs desktop by title content
                    if url and url.startswith("http"):
                        self._record_web_target(url, title, "", display_name, filename, local)
                    else:
                        app_exe = Path(app).name if app else ""
                        self._record_window_target(app_exe, title, display_name, filename, local)
                if app and not sel and not title:
                    app_exe = Path(app).name if app else ""
                    self._record_window_target(app_exe, "", display_name, filename, local)

    # ── Deep selector parser ───────────────────────────────────────
    def _deep_parse_selector(self, selector_raw: str, activity: str, xaml_file: str, activity_type: str = ""):
        """
        Parse UiPath selector XML (possibly HTML-entity-encoded) and extract:
          - <html app='...' title='...' url='...'>  → web targets
          - <wnd app='...' title='...'>             → window targets
          - <ctrl ...>, <webctrl ...>               → supplemental info
        """
        # Decode HTML entities that UiPath uses in selector attributes
        sel = selector_raw.replace("&lt;","<").replace("&gt;",">").replace("&amp;","&").replace("&quot;",'"')

        # Wrap in a root so ET can parse it
        try:
            root = ET.fromstring(f"<root>{sel}</root>")
        except ET.ParseError:
            # Fallback: regex scan
            self._regex_selector_fallback(sel, activity, xaml_file)
            return

        app_exe    = ""
        window_title = ""
        browser    = ""
        page_title = ""
        url        = ""

        for node in root:
            tag = node.tag.lower()
            attrs = {k.lower(): v for k, v in node.attrib.items()}

            if tag == "wnd":
                a = attrs.get("app","")
                t = attrs.get("title","")
                if a: app_exe = a
                if t: window_title = t

            elif tag in ("html","browser","web"):
                a = attrs.get("app","")
                t = attrs.get("title","")
                u = attrs.get("url","") or attrs.get("urlpart","")
                if a: browser = a
                if t: page_title = t
                if u and u.startswith("http"): url = u

            elif tag == "webctrl":
                # webctrl can sometimes carry a url
                u = attrs.get("url","") or attrs.get("urlpart","")
                if u and u.startswith("http") and not url:
                    url = u

        # Record web target if we found browser/page info
        if browser or page_title or url:
            self._record_web_target(url, page_title, browser, activity, xaml_file)

        # Record window target if we found wnd info
        if app_exe or window_title:
            self._record_window_target(app_exe, window_title, activity, xaml_file)

    def _regex_selector_fallback(self, sel: str, activity: str, xaml_file: str):
        """Regex-based selector scan when XML parse fails."""
        # Window app/title
        wnd_match = re.search(r"<wnd[^>]*app='([^']*)'[^>]*title='([^']*)'", sel, re.I)
        if not wnd_match:
            wnd_match = re.search(r"<wnd[^>]*title='([^']*)'[^>]*app='([^']*)'", sel, re.I)
        if wnd_match:
            self._record_window_target(wnd_match.group(1), wnd_match.group(2), activity, xaml_file)

        # Browser/HTML
        html_match = re.search(r"<html[^>]*app='([^']*)'[^>]*(?:title='([^']*)')?[^>]*(?:url='([^']*)')?", sel, re.I)
        if html_match:
            self._record_web_target(
                html_match.group(3) or "", html_match.group(2) or "",
                html_match.group(1) or "", activity, xaml_file
            )

        # Loose URL scan
        for url in re.findall(r"https?://[^\s'\"<>]+", sel):
            self._record_web_target(url, "", "", activity, xaml_file)

    def _guess_browser(self, elem) -> str:
        """Guess browser from BrowserType attribute."""
        bt = elem.get("BrowserType","") or elem.get("browserType","")
        mapping = {"Chrome":"chrome.exe","Firefox":"firefox.exe",
                   "Edge":"msedge.exe","IE":"iexplore.exe","InternetExplorer":"iexplore.exe"}
        return mapping.get(bt, bt)

    def _record_web_target(self, url, page_title, browser, activity, xaml_file,
                            activity_type: str = ""):
        """Deduplicated append to web_targets."""
        # ── Normalise browser name ────────────────────────────────
        browser_map = {
            "chrome.exe":"Google Chrome","firefox.exe":"Mozilla Firefox",
            "msedge.exe":"Microsoft Edge","iexplore.exe":"Internet Explorer",
            "chrome":"Google Chrome","firefox":"Mozilla Firefox",
            "edge":"Microsoft Edge","ie":"Internet Explorer",
        }
        browser_friendly = browser_map.get((browser or "").lower(), browser) if browser else ""

        is_modern = activity_type in (
            "UseApplicationBrowser","UseBrowser","UseApplication",
            "BrowserScope","DesktopScope","ApplicationCard","NavigateTo",
            "NavigateToActivity","BrowserNavigateTo",
        )

        url_norm  = url.rstrip("/") if url else ""
        page_norm = page_title.strip() if page_title else ""

        # ── Smart dedup: merge records with same URL or same page title ──
        # If a record exists with the same URL (non-empty), enrich it rather
        # than adding a duplicate. Same for matching page_title.
        for existing in self.results["web_targets"]:
            eu = existing["url"].rstrip("/") if existing["url"] else ""
            et = existing["page_title"].strip() if existing["page_title"] else ""
            # Match on URL
            if url_norm and eu == url_norm:
                if page_norm and not et:
                    existing["page_title"] = page_title
                if browser_friendly and not existing["browser"]:
                    existing["browser"] = browser_friendly
                if url and url.startswith("http"):
                    self.results["web_urls"].add(url)
                return
            # Match on page title (within same xaml file)
            if page_norm and et == page_norm and existing["xaml_file"] == xaml_file:
                if url_norm and not eu:
                    existing["url"] = url
                if browser_friendly and not existing["browser"]:
                    existing["browser"] = browser_friendly
                if url and url.startswith("http"):
                    self.results["web_urls"].add(url)
                return

        # ── New record ────────────────────────────────────────────
        if url_norm or page_norm:
            self.results["web_targets"].append({
                "url":            url,
                "page_title":     page_title,
                "browser":        browser_friendly,
                "activity":       activity,
                "xaml_file":      xaml_file,
                "activity_model": "Modern" if is_modern else "Classic",
            })
        if url and url.startswith("http"):
            self.results["web_urls"].add(url)

    def _record_window_target(self, app_exe, window_title, activity, xaml_file,
                               activity_type: str = ""):
        """Deduplicated append to window_targets."""
        # ── Always strip full path to basename first ──────────────
        # Use PureWindowsPath to handle Windows-style paths on any OS
        if app_exe and ("\\" in app_exe or "/" in app_exe or os.sep in app_exe):
            from pathlib import PureWindowsPath, PurePosixPath
            try:
                app_exe = PureWindowsPath(app_exe).name or PurePosixPath(app_exe).name or app_exe
            except Exception:
                app_exe = app_exe.replace("\\","/").split("/")[-1] or app_exe

        exe_friendly_map = {
            "saplogon.exe":        "SAP Logon / SAP GUI",
            "payrolltool.exe":     "Payroll Management System",
            "financeapproval.exe": "Finance Approval System",
            "excel.exe":           "Microsoft Excel",
            "winword.exe":         "Microsoft Word",
            "powerpnt.exe":        "Microsoft PowerPoint",
            "outlook.exe":         "Microsoft Outlook",
            "notepad.exe":         "Notepad",
            "powershell.exe":      "PowerShell",
            "cmd.exe":             "Command Prompt",
            "acrobat.exe":         "Adobe Acrobat",
            "acrord32.exe":        "Adobe Acrobat Reader",
        }
        app_friendly = exe_friendly_map.get((app_exe or "").lower(), app_exe) if app_exe else ""

        # ── Guard: skip if this looks like a web/browser target ──
        # Browser EXE names should never appear in window_targets
        BROWSER_EXES = {"chrome.exe","msedge.exe","firefox.exe","iexplore.exe","chromium.exe"}
        if (app_exe or "").lower() in BROWSER_EXES:
            return
        # Titles containing known web app keywords with no desktop EXE → likely web
        BROWSER_TITLE_CLUES = ["workday","salesforce","servicenow","sharepoint"]
        if not app_exe and window_title:
            tl = window_title.lower()
            if any(k in tl for k in BROWSER_TITLE_CLUES):
                return

        # ── Friendly name fallback from title keywords ─────────────
        if not app_friendly or app_friendly == app_exe:
            title_lower = (window_title or "").lower()
            for pattern, friendly in WINDOWS_APP_PATTERNS.items():
                if re.search(pattern, title_lower) or (app_exe and re.search(pattern, app_exe.lower())):
                    app_friendly = friendly
                    break

        is_modern = activity_type in (
            "UseApplication","ApplicationCard","DesktopScope","UseApplicationBrowser"
        )
        # ── Dedup on (exe_basename, window_title) ─────────────────
        # Normalise: strip any full path to basename for key comparison
        exe_key = Path(app_exe).name.lower() if app_exe else ""
        key = (exe_key, window_title or "")
        existing_keys = {
            (Path(t["app_exe"]).name.lower() if t["app_exe"] else "", t["window_title"] or "")
            for t in self.results["window_targets"]
        }
        # Always compute app_exe_store (basename) before the dedup block
        app_exe_store = Path(app_exe).name if (app_exe and ("\\" in app_exe or "/" in app_exe or os.sep in app_exe)) else (app_exe or "")

        if key not in existing_keys and (app_exe or window_title):
            self.results["window_targets"].append({
                "app_exe":        app_exe_store,
                "app_friendly":   app_friendly or "",
                "window_title":   window_title or "",
                "activity":       activity,
                "xaml_file":      xaml_file,
                "activity_model": "Modern" if is_modern else "Classic",
            })
        if app_friendly:
            self.results["applications"].add(app_friendly)
            self.results["window_selectors"].add(app_friendly)
        elif app_exe_store:
            # Only add basename exe if no friendly name — avoid full paths in applications list
            self.results["applications"].add(app_exe_store)

    def _classify_activity(self, local_name, elem):
        # ── Skip structural/property wrapper nodes ────────────────────
        NOISE_SUFFIXES = (".Target",".Url",".URL",".Selector",".Text",".Input",
                          ".Output",".Properties",".Anchor",".Timeout",".Options",
                          ".Condition",".Body",".Then",".Else",".Finally",".Catches",
                          ".Variables",".Arguments",".Imports",".Members",
                          ".TargetApp",".IfExists",".IfNotExists")
        NOISE_EXACT    = {"Target","Selector","Anchor","TargetAnchor","TargetAnchorable",
                          "TargetApp","InArgument","OutArgument","InOutArgument",
                          "WindowControlDescriptor","WebControlDescriptor",
                          "ControlDescriptor","ActivityAction","ActivityFunc",
                          "Variable","Argument","DelegateInArgument",
                          "DelegateOutArgument","TextExpression",
                          "VisualBasicValue","VisualBasicReference",
                          "CSharpValue","CSharpReference"}
        if any(local_name.endswith(s) for s in NOISE_SUFFIXES):
            return (None, None)
        if local_name in NOISE_EXACT:
            return (None, None)

        if local_name in ACTIVITY_CATALOG:
            return ACTIVITY_CATALOG[local_name]

        lower = local_name.lower()

        # ── N-prefixed modern UI Automation activities ────────────────
        if local_name == "NApplicationCard":
            return ("UI Automation [Modern]", "Use Application/Browser")
        if local_name == "NGetText":
            return ("UI Automation [Modern]", "Get Text")
        if local_name == "NClick":
            return ("UI Automation [Modern]", "Click")
        if local_name == "NTypeInto":
            return ("UI Automation [Modern]", "Type Into")
        if local_name == "NCheckState":
            return ("UI Automation [Modern]", "Check App State")
        if local_name == "NKeyboardShortcuts":
            return ("UI Automation [Modern]", "Keyboard Shortcuts")
        if local_name == "NNavigateBrowser":
            return ("Web Browser", "Navigate Browser")
        if local_name.startswith("N") and any(k in lower for k in
           ["click","type","get","check","hover","select","drag","scroll","key","navigate"]):
            return ("UI Automation [Modern]", local_name)

        # ── Modern Excel activities ───────────────────────────────────
        if local_name in ("ExcelApplicationCard",):
            return ("Excel", "Use Excel File (Modern)")
        if local_name in ("ExcelProcessScopeX",):
            return ("Excel", "Excel Process Scope (Modern)")
        if local_name in ("ExcelForEachRowX",):
            return ("Excel", "For Each Excel Row (Modern)")
        if local_name in ("WriteRangeX",):
            return ("Excel", "Write DataTable to Excel (Modern)")
        if local_name in ("ReadRangeX",):
            return ("Excel", "Read Range (Modern)")

        if any(k in lower for k in ["excel","workbook","sheet","range","cell"]):
            return ("Excel", local_name)
        if any(k in lower for k in ["browser","web","navigate","url","http"]):
            return ("Web Browser", local_name)
        if any(k in lower for k in ["click","type","hotkey","attach","find","element","window"]):
            return ("UI Automation", local_name)
        if any(k in lower for k in ["mail","email","outlook","smtp","imap"]):
            return ("Outlook/Email", local_name)
        if any(k in lower for k in ["pdf","ocr"]):
            return ("PDF", local_name)
        if any(k in lower for k in ["queue","transaction","asset","credential"]):
            return ("Orchestrator", local_name)
        if any(k in lower for k in ["sap"]):
            return ("SAP", local_name)
        if any(k in lower for k in ["database","sql","query","executenonquery","connect"]):
            return ("Database", local_name)
        if any(k in lower for k in ["log","assign","invoke","delay","file","folder","process","zip"]):
            return ("System", local_name)
        return (None, None)

    def extract_project_info(self):
        pj = self.project_json
        self.results["project_info"] = {
            "name": pj.get("name","N/A"),
            "description": pj.get("description","N/A"),
            "main_workflow": pj.get("main","N/A"),
            "project_version": pj.get("projectVersion","N/A"),
            "dependencies": pj.get("dependencies",{}),
            "analyzed_on": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "path": str(self.project_path),
        }

    def build_summary(self):
        acts = self.results["activities"]
        self.results["summary"] = {
            "total_activities_unique": len(acts),
            "total_activity_usages": sum(a["count"] for a in acts),
            "total_xaml_files": len(self.xaml_files),
            "applications_detected": len(self.results["applications"]),
            "web_urls_found": len(self.results["web_urls"]),
            "web_targets_found": len(self.results["web_targets"]),
            "window_targets_found": len(self.results["window_targets"]),
            "invoked_workflows": len(self.results["invoked_workflows"]),
            "categories": sorted(set(a["category"] for a in acts)),
        }

    def run(self):
        if not self.discover():
            return False
        self.parse_all_xaml()
        self.extract_project_info()
        self.build_summary()
        return True


# ══════════════════════════════════════════════════════════
#  Reusability Engine  (rule-based, fully offline)
# ══════════════════════════════════════════════════════════
class ReusabilityEngine:
    def __init__(self, projects: list):
        """projects: list of UiPathProjectAnalyzer (already run)"""
        self.projects = projects

    # ── Workflow-level matrix ──────────────────────────────
    def workflow_matrix(self) -> list:
        """
        Returns one row per unique workflow filename (basename),
        scored and described.
        """
        # Aggregate: workflow_name → {projects, total_usages, categories, invoked_by}
        wf_agg = defaultdict(lambda: {
            "projects": set(), "total_invocations": 0,
            "categories": set(), "invoked_by_projects": set()
        })

        for proj in self.projects:
            pname = proj.results["project_info"]["name"]
            for wf_detail in proj.results["workflow_details"]:
                fname = Path(wf_detail["file"]).name
                wf_agg[fname]["projects"].add(pname)
                for cat in wf_detail["categories"]:
                    wf_agg[fname]["categories"].add(cat)
            for inv in proj.results["invoked_workflows"]:
                wf_name = Path(inv["workflow"]).name
                wf_agg[wf_name]["total_invocations"] += 1
                wf_agg[wf_name]["invoked_by_projects"].add(pname)

        rows = []
        total_projects = len(self.projects)

        for wf_name, data in sorted(wf_agg.items()):
            score, description = self._score_workflow(
                wf_name, data, total_projects
            )
            shared_candidate = score >= SHARED_LIB_THRESHOLD
            rows.append({
                "workflow_file":        wf_name,
                "problem_solved":       description,
                "projects_containing":  len(data["projects"]),
                "projects_list":        ", ".join(sorted(data["projects"])),
                "total_invocations":    data["total_invocations"],
                "activity_categories":  ", ".join(sorted(data["categories"])),
                "reusability_score":    score,
                "score_label":          self._score_label(score),
                "shared_lib_candidate": "✅ YES" if shared_candidate else "—",
                "recommendation":       self._recommend_workflow(score, data, wf_name),
            })

        rows.sort(key=lambda x: -x["reusability_score"])
        return rows

    def _score_workflow(self, wf_name, data, total_projects):
        name_lower = wf_name.lower().replace("_","").replace("-","").replace(".xaml","")
        base_score = 40

        # +20 pts: cross-project presence
        proj_ratio = len(data["projects"]) / max(total_projects, 1)
        base_score += int(proj_ratio * 20)

        # +15 pts: invocation frequency
        inv_score = min(data["total_invocations"] * 5, 15)
        base_score += inv_score

        # +up to 20 pts: workflow name pattern match
        description = f"Workflow file used in {len(data['projects'])} project(s)."
        pattern_boost = 0
        for pattern, desc, boost in WORKFLOW_PATTERNS:
            if re.search(pattern, name_lower):
                description = desc
                pattern_boost = boost
                break
        base_score += pattern_boost

        # category bonus: high-value categories
        for cat in data["categories"]:
            cr = CATEGORY_REUSABILITY.get(cat, (50,""))[0]
            if cr >= 80:
                base_score += 3  # bonus per high-value category

        return min(base_score, 100), description

    def _score_label(self, score):
        if score >= 80: return "🟢 High"
        if score >= 60: return "🟡 Medium"
        if score >= 40: return "🟠 Low-Medium"
        return "🔴 Low"

    def _recommend_workflow(self, score, data, wf_name):
        if score >= 80 and len(data["projects"]) >= 2:
            return "Extract to shared reusable library. Standardise interface (arguments) and publish to Orchestrator."
        if score >= 60 and len(data["projects"]) >= 2:
            return "Good candidate for a shared component. Parameterise and document before sharing."
        if score >= 60:
            return "Consider making this reusable — abstract hardcoded values to arguments."
        if data["total_invocations"] == 0:
            return "Not invoked by other workflows. Review if it can be merged or generalised."
        return "Keep project-specific for now. Monitor usage as more projects are added."

    # ── Activity-category level matrix ────────────────────
    def activity_matrix(self) -> list:
        cat_agg = defaultdict(lambda: {
            "total_usages": 0, "projects_using": set(),
            "unique_activity_types": set()
        })
        for proj in self.projects:
            pname = proj.results["project_info"]["name"]
            for act in proj.results["activities"]:
                cat = act["category"]
                cat_agg[cat]["total_usages"]         += act["count"]
                cat_agg[cat]["projects_using"].add(pname)
                cat_agg[cat]["unique_activity_types"].add(act["raw_tag"])

        rows = []
        total_projects = len(self.projects)
        for cat, data in sorted(cat_agg.items()):
            base_score, cat_desc = CATEGORY_REUSABILITY.get(cat, (50,"General automation activities."))
            # Adjust score by cross-project spread
            spread = len(data["projects_using"]) / max(total_projects, 1)
            adjusted = min(int(base_score * 0.7 + spread * 30), 100)
            rows.append({
                "category":              cat,
                "description":           ACTIVITY_DESCRIPTIONS.get(cat, cat_desc),
                "unique_activity_types": len(data["unique_activity_types"]),
                "total_usages":          data["total_usages"],
                "projects_using":        len(data["projects_using"]),
                "projects_list":         ", ".join(sorted(data["projects_using"])),
                "reusability_score":     adjusted,
                "score_label":           self._score_label(adjusted),
                "shared_lib_candidate":  "✅ YES" if adjusted >= SHARED_LIB_THRESHOLD else "—",
                "recommendation":        self._recommend_category(cat, adjusted, data, total_projects),
            })
        rows.sort(key=lambda x: -x["reusability_score"])
        return rows

    def _recommend_category(self, cat, score, data, total_projects):
        spread = len(data["projects_using"])
        if score >= 80 and spread >= 2:
            return f"Create a '{cat} Utilities' shared library with standardised wrappers."
        if score >= 70:
            return f"Document and template common {cat} patterns for team reuse."
        if cat == "UI Automation" and score < 60:
            return "Selector-dependent — reuse within same application family only."
        if cat == "Control Flow":
            return "Extract TryCatch and exception handling into a standard RE Framework template."
        return "Monitor growth; extract shared patterns when 2+ projects share identical logic."

    # ── Shared Library Recommendations ────────────────────
    def shared_library_recommendations(self, wf_matrix, act_matrix) -> list:
        recs = []

        # From high-scoring workflows used in multiple projects
        seen = set()
        for row in wf_matrix:
            if row["reusability_score"] >= SHARED_LIB_THRESHOLD and row["projects_containing"] >= 2:
                key = row["workflow_file"]
                if key not in seen:
                    seen.add(key)
                    recs.append({
                        "type":           "Workflow",
                        "component":      row["workflow_file"],
                        "score":          row["reusability_score"],
                        "used_in":        row["projects_list"],
                        "problem_solved": row["problem_solved"],
                        "action":         "Extract to shared .xaml, define standard In/Out arguments, publish as NuGet package.",
                    })

        # From high-scoring categories
        for row in act_matrix:
            if row["reusability_score"] >= 80 and row["projects_using"] >= 2:
                recs.append({
                    "type":           "Activity Category",
                    "component":      f"{row['category']} Utilities",
                    "score":          row["reusability_score"],
                    "used_in":        row["projects_list"],
                    "problem_solved": row["description"],
                    "action":         row["recommendation"],
                })

        recs.sort(key=lambda x: -x["score"])
        return recs


# ══════════════════════════════════════════════════════════
#  Repo Discovery
# ══════════════════════════════════════════════════════════
def discover_projects(root_path: Path) -> list:
    """
    Walk directory tree and find all UiPath project roots.
    A project root is a folder that contains a project.json AND
    at least one .xaml file (directly or in subfolders).

    Strategy:
      1. Collect every folder that has a project.json.
      2. Remove any folder that is an ancestor of another — this prevents
         a workspace/solution-level project.json from swallowing all the
         individual projects inside it.
      3. If only one candidate remains (the root itself) AND it has
         sub-projects inside, prefer the sub-projects.
    """
    root_path = Path(root_path).resolve()

    # Find every folder containing a project.json
    candidates = sorted({pjson.parent.resolve() for pjson in root_path.rglob("project.json")})

    if not candidates:
        return []

    # Remove ancestor folders: if folder A contains folder B (both are candidates),
    # keep only B (the more specific project), unless A has its own XAML files that
    # are NOT inside B.
    def is_ancestor_of_any(folder, others):
        """True if folder is a parent of at least one other candidate."""
        return any(other != folder and str(other).startswith(str(folder) + os.sep)
                   for other in others)

    # Filter: drop any candidate that is a pure ancestor of another candidate
    # (i.e. it only exists because of nested sub-projects)
    filtered = []
    for c in candidates:
        if is_ancestor_of_any(c, candidates):
            # Only keep this ancestor if it has its own XAML files not inside sub-projects
            sub_project_paths = [o for o in candidates if o != c and str(o).startswith(str(c) + os.sep)]
            own_xamls = [
                x for x in c.glob("*.xaml")  # direct children only
                if not any(str(x).startswith(str(sp) + os.sep) for sp in sub_project_paths)
            ]
            if own_xamls:
                filtered.append(c)
            # else: skip — it's just a container folder
        else:
            filtered.append(c)

    return sorted(filtered) if filtered else candidates


# ══════════════════════════════════════════════════════════
#  Console Reporting
# ══════════════════════════════════════════════════════════
def print_project_report(analyzer: UiPathProjectAnalyzer):
    pi  = analyzer.results["project_info"]
    sm  = analyzer.results["summary"]
    acts= analyzer.results["activities"]
    apps= sorted(analyzer.results["applications"])
    web_targets = analyzer.results["web_targets"]
    win_targets = analyzer.results["window_targets"]

    print(f"\n{C.CYAN}{'═'*65}{C.RESET}")
    print(f"{C.BOLD}{C.WHITE}  {pi['name']}  —  Project Report{C.RESET}")
    print(f"{C.CYAN}{'═'*65}{C.RESET}")
    print(f"  {C.BOLD}Description:{C.RESET}  {pi['description']}")
    print(f"  {C.BOLD}Version:{C.RESET}      {pi['project_version']}   {C.GREY}|{C.RESET}  {C.BOLD}Main:{C.RESET} {pi['main_workflow']}")
    print(f"\n  XAML files: {sm['total_xaml_files']}  |  Unique activities: {sm['total_activities_unique']}  |  "
          f"Total usages: {sm['total_activity_usages']}  |  Apps detected: {sm['applications_detected']}")

    current_cat = None
    for act in acts:
        if act["category"] != current_cat:
            current_cat = act["category"]
            print(f"\n  {C.MAGENTA}{C.BOLD}▶ {current_cat}{C.RESET}")
            print(f"  {'Activity':<38} {'Count':>5}")
            print(f"  {'─'*38} {'─'*5}")
        name = act["friendly_name"] or act["raw_tag"]
        print(f"  {name:<38} {act['count']:>5}")

    print(f"\n  {C.BOLD}Applications:{C.RESET}")
    for app in apps:
        icon = "🌐" if any(k in app.lower() for k in ["browser","chrome","firefox","edge","web"]) else "🖥️"
        print(f"    {icon}  {app}")

    # ── Web Targets ──────────────────────────────────────────
    if web_targets:
        print(f"\n  {C.BOLD}{C.CYAN}🌐 Web Browser Targets ({len(web_targets)} unique):{C.RESET}")
        print(f"  {'URL':<50} {'Page Title':<35} {'Browser':<18} {'Activity'}")
        print(f"  {'─'*50} {'─'*35} {'─'*18} {'─'*25}")
        for t in web_targets:
            url   = (t["url"][:48]+"..") if len(t["url"])>50 else t["url"]
            title = (t["page_title"][:33]+"..") if len(t["page_title"])>35 else t["page_title"]
            brow  = t["browser"] or "—"
            act   = t["activity"][:24] if t["activity"] else "—"
            print(f"  {C.BLUE}{url:<50}{C.RESET} {title:<35} {brow:<18} {C.GREY}{act}{C.RESET}")

    # ── Window Targets ───────────────────────────────────────
    if win_targets:
        print(f"\n  {C.BOLD}{C.CYAN}🖥️  Windows Desktop Targets ({len(win_targets)} unique):{C.RESET}")
        print(f"  {'Process (EXE)':<28} {'Window Title':<40} {'Activity'}")
        print(f"  {'─'*28} {'─'*40} {'─'*25}")
        for t in win_targets:
            exe   = t["app_exe"] or "—"
            title = (t["window_title"][:38]+"..") if len(t["window_title"])>40 else t["window_title"]
            act   = t["activity"][:24] if t["activity"] else "—"
            print(f"  {C.YELLOW}{exe:<28}{C.RESET} {title:<40} {C.GREY}{act}{C.RESET}")

def print_reusability_report(wf_matrix, act_matrix, shared_libs, total_projects):
    print(f"\n{C.CYAN}{'═'*70}{C.RESET}")
    print(f"{C.BOLD}{C.WHITE}  REUSABILITY MATRIX  ({total_projects} projects scanned){C.RESET}")
    print(f"{C.CYAN}{'═'*70}{C.RESET}")

    print(f"\n{C.BOLD}{C.YELLOW}  ▶ WORKFLOW-LEVEL REUSABILITY{C.RESET}")
    print(f"  {'Workflow':<28} {'Score':>5}  {'Label':<14} {'Projects':>8}  {'Invocations':>11}  {'Shared Lib?'}")
    print(f"  {'─'*28} {'─'*5}  {'─'*14} {'─'*8}  {'─'*11}  {'─'*11}")
    for r in wf_matrix:
        print(f"  {r['workflow_file']:<28} {r['reusability_score']:>5}  {r['score_label']:<14} "
              f"{r['projects_containing']:>8}  {r['total_invocations']:>11}  {r['shared_lib_candidate']}")
        print(f"    {C.GREY}↳ {r['problem_solved']}{C.RESET}")
        print(f"    {C.GREY}  💡 {r['recommendation']}{C.RESET}")

    print(f"\n{C.BOLD}{C.YELLOW}  ▶ ACTIVITY CATEGORY REUSABILITY{C.RESET}")
    print(f"  {'Category':<24} {'Score':>5}  {'Label':<14} {'Projects':>8}  {'Usages':>7}  {'Shared Lib?'}")
    print(f"  {'─'*24} {'─'*5}  {'─'*14} {'─'*8}  {'─'*7}  {'─'*11}")
    for r in act_matrix:
        print(f"  {r['category']:<24} {r['reusability_score']:>5}  {r['score_label']:<14} "
              f"{r['projects_using']:>8}  {r['total_usages']:>7}  {r['shared_lib_candidate']}")
        print(f"    {C.GREY}↳ {r['description']}{C.RESET}")

    if shared_libs:
        print(f"\n{C.BOLD}{C.GREEN}  ▶ SHARED LIBRARY CANDIDATES  ({len(shared_libs)} identified){C.RESET}")
        for i, lib in enumerate(shared_libs, 1):
            print(f"\n  {C.BOLD}{i}. [{lib['type']}] {lib['component']}  (score: {lib['score']}){C.RESET}")
            print(f"     Used in:  {lib['used_in']}")
            print(f"     Solves:   {lib['problem_solved']}")
            print(f"     Action:   {C.CYAN}{lib['action']}{C.RESET}")

    print(f"\n{C.CYAN}{'═'*70}{C.RESET}\n")


# ══════════════════════════════════════════════════════════
#  Excel Export  (full multi-sheet workbook)
# ══════════════════════════════════════════════════════════
def export_excel(projects, wf_matrix, act_matrix, shared_libs, output_dir: Path) -> Path:
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    filename = output_dir / f"{label}_UiPath_Report_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Styles ──
    HDR   = PatternFill("solid", fgColor="1F3864")
    CAT   = PatternFill("solid", fgColor="2E75B6")
    ALT   = PatternFill("solid", fgColor="EBF3FB")
    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
    AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")
    RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
    HDR_F = Font(bold=True, color="FFFFFF", size=11)
    CAT_F = Font(bold=True, color="FFFFFF", size=10)
    TITLE_F = Font(bold=True, size=14, color="1F3864")
    thin  = Side(style="thin", color="BDD7EE")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)

    def hdr(ws, row, cols, fill=HDR, font=HDR_F):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill=fill; cell.font=font
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=bdr
        ws.row_dimensions[row].height=22

    def autowidth(ws, extra=4):
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+extra, 55)

    def score_fill(score):
        if score >= 80: return GREEN_FILL
        if score >= 60: return AMBER_FILL
        return RED_FILL

    # ─────────────────────────────────────────────────
    # Sheet 1 — Repo Summary
    # ─────────────────────────────────────────────────
    ws = wb.active; ws.title = "📊 Repo Summary"
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:D1")
    t=ws["A1"]; t.value="UiPath RPA Repository Analysis"
    t.font=TITLE_F; t.alignment=Alignment(horizontal="center")
    ws.append([])
    ws.append(["Analyzed On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total Projects", len(projects)])
    ws.append(["Total XAML Files", sum(len(p.xaml_files) for p in projects)])
    ws.append(["Total Unique Activity Types",
               len(set(a["raw_tag"] for p in projects for a in p.results["activities"]))])
    ws.append(["Total Activity Usages",
               sum(a["count"] for p in projects for a in p.results["activities"])])
    ws.append([])
    ws.append(["#","Project","Description","Version","XAML Files","Activities","Applications"])
    hdr(ws, ws.max_row, ["#","Project","Description","Version","XAML Files","Activities","Applications"])
    for i, proj in enumerate(projects, 1):
        pi=proj.results["project_info"]; sm=proj.results["summary"]
        ws.append([i, pi["name"], pi["description"], pi["project_version"],
                   sm["total_xaml_files"], sm["total_activity_usages"],
                   sm["applications_detected"]])
        r=ws.max_row
        fill=ALT if i%2==0 else PatternFill()
        for c in range(1,8):
            cell=ws.cell(r,c)
            if fill.patternType: cell.fill=fill
            cell.border=bdr
    autowidth(ws)

    # ─────────────────────────────────────────────────
    # Sheet 2 — Workflow Reusability Matrix
    # ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("🔁 Workflow Matrix")
    ws2.sheet_view.showGridLines=False
    cols=["Workflow File","Problem Solved","Projects Containing","Projects List",
          "Total Invocations","Activity Categories","Reusability Score","Score Label",
          "Shared Lib Candidate","Recommendation"]
    hdr(ws2,1,cols)
    for i, row in enumerate(wf_matrix):
        ws2.append([
            row["workflow_file"], row["problem_solved"],
            row["projects_containing"], row["projects_list"],
            row["total_invocations"], row["activity_categories"],
            row["reusability_score"], row["score_label"],
            row["shared_lib_candidate"], row["recommendation"]
        ])
        r=ws2.max_row
        sf=score_fill(row["reusability_score"])
        ws2.cell(r,7).fill=sf; ws2.cell(r,8).fill=sf
        for c in range(1,11):
            ws2.cell(r,c).border=bdr
            ws2.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
        ws2.row_dimensions[r].height=40
    autowidth(ws2)

    # ─────────────────────────────────────────────────
    # Sheet 3 — Activity Category Matrix
    # ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("📦 Activity Matrix")
    ws3.sheet_view.showGridLines=False
    cols=["Category","Description","Unique Activity Types","Total Usages",
          "Projects Using","Projects List","Reusability Score","Score Label",
          "Shared Lib Candidate","Recommendation"]
    hdr(ws3,1,cols)
    for i, row in enumerate(act_matrix):
        ws3.append([
            row["category"], row["description"],
            row["unique_activity_types"], row["total_usages"],
            row["projects_using"], row["projects_list"],
            row["reusability_score"], row["score_label"],
            row["shared_lib_candidate"], row["recommendation"]
        ])
        r=ws3.max_row
        sf=score_fill(row["reusability_score"])
        ws3.cell(r,7).fill=sf; ws3.cell(r,8).fill=sf
        for c in range(1,11):
            ws3.cell(r,c).border=bdr
            ws3.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
        ws3.row_dimensions[r].height=40
    autowidth(ws3)

    # ─────────────────────────────────────────────────
    # Sheet 4 — Shared Library Candidates
    # ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("⭐ Shared Library Candidates")
    ws4.sheet_view.showGridLines=False
    if shared_libs:
        hdr(ws4,1,["#","Type","Component","Score","Used In Projects","Problem Solved","Recommended Action"])
        for i, lib in enumerate(shared_libs, 1):
            ws4.append([i, lib["type"], lib["component"], lib["score"],
                        lib["used_in"], lib["problem_solved"], lib["action"]])
            r=ws4.max_row
            ws4.cell(r,4).fill=score_fill(lib["score"])
            for c in range(1,8):
                ws4.cell(r,c).border=bdr
                ws4.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
            ws4.row_dimensions[r].height=45
    else:
        ws4["A1"]="No shared library candidates identified yet. Analyse more projects to find patterns."
    autowidth(ws4)

    # ─────────────────────────────────────────────────
    # Sheet 5 — All Activities (cross-project)
    # ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("📋 All Activities")
    ws5.sheet_view.showGridLines=False
    hdr(ws5,1,["Project","Category","Activity Name","Raw Tag","Usage Count","XAML Files"])
    row_num=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for act in proj.results["activities"]:
            files_str=", ".join(Path(f).name for f in act["files"])
            ws5.append([pname, act["category"],
                        act["friendly_name"] or act["raw_tag"],
                        act["raw_tag"], act["count"], files_str])
            r=ws5.max_row; row_num+=1
            fill=ALT if row_num%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws5.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
    autowidth(ws5)

    # ─────────────────────────────────────────────────
    # Sheet 6 — Applications Map
    # ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("🖥️ Applications Map")
    ws6.sheet_view.showGridLines=False
    hdr(ws6,1,["Application","Type","Projects Using It"])
    app_proj = defaultdict(set)
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for app in proj.results["applications"]:
            app_proj[app].add(pname)
    for i, (app, projs) in enumerate(sorted(app_proj.items()), 1):
        atype = "Web" if any(k in app.lower() for k in ["browser","chrome","firefox","edge"]) \
                else "Windows Desktop" if any(k in app.lower() for k in ["excel","word","outlook","sap"]) \
                else "Platform / Service"
        ws6.append([app, atype, ", ".join(sorted(projs))])
        r=ws6.max_row
        fill=ALT if i%2==0 else PatternFill()
        for c in range(1,4):
            cell=ws6.cell(r,c)
            if fill.patternType: cell.fill=fill
            cell.border=bdr
    autowidth(ws6)

    # ─────────────────────────────────────────────────
    # Sheet 7 — Dependencies cross-ref
    # ─────────────────────────────────────────────────
    ws7 = wb.create_sheet("📦 Dependencies")
    ws7.sheet_view.showGridLines=False
    hdr(ws7,1,["Package","Version","Projects Using"])
    pkg_proj = defaultdict(lambda: defaultdict(set))
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for pkg, ver in proj.results["project_info"]["dependencies"].items():
            pkg_proj[pkg][ver].add(pname)
    for i, (pkg, versions) in enumerate(sorted(pkg_proj.items()), 1):
        for ver, projs in versions.items():
            ws7.append([pkg, ver, ", ".join(sorted(projs))])
            r=ws7.max_row
            fill=ALT if i%2==0 else PatternFill()
            for c in range(1,4):
                cell=ws7.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
    autowidth(ws7)

    # ─────────────────────────────────────────────────
    # Sheet 8 — Web Browser Targets (URLs + Page Titles)
    # ─────────────────────────────────────────────────
    ws8 = wb.create_sheet("🌐 Web Targets")
    ws8.sheet_view.showGridLines=False
    hdr(ws8,1,["Project","URL","Page Title","Browser","Activity","XAML File"])
    row_n=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["web_targets"]:
            ws8.append([pname, t["url"], t["page_title"],
                        t["browser"] or "—", t["activity"],
                        Path(t["xaml_file"]).name])
            r=ws8.max_row; row_n+=1
            fill=ALT if row_n%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws8.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
                cell.alignment=Alignment(vertical="center",wrap_text=True)
            # Hyperlink the URL cell
            url_val=t["url"]
            if url_val and url_val.startswith("http"):
                ws8.cell(r,2).hyperlink=url_val
                ws8.cell(r,2).font=Font(color="0563C1",underline="single")
    if row_n==0:
        ws8["A2"]="No web browser targets found. Ensure XAML files contain OpenBrowser/AttachBrowser/Navigate activities with Selector or Url attributes."
    autowidth(ws8)

    # ─────────────────────────────────────────────────
    # Sheet 9 — Windows Desktop Targets
    # ─────────────────────────────────────────────────
    ws9 = wb.create_sheet("🖥️ Window Targets")
    ws9.sheet_view.showGridLines=False
    hdr(ws9,1,["Project","Process (EXE)","Friendly App Name","Window Title","Activity","XAML File"])
    row_n=0
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["window_targets"]:
            ws9.append([pname, t["app_exe"], t["app_friendly"],
                        t["window_title"], t["activity"],
                        Path(t["xaml_file"]).name])
            r=ws9.max_row; row_n+=1
            fill=ALT if row_n%2==0 else PatternFill()
            for c in range(1,7):
                cell=ws9.cell(r,c)
                if fill.patternType: cell.fill=fill
                cell.border=bdr
                cell.alignment=Alignment(vertical="center",wrap_text=True)
    if row_n==0:
        ws9["A2"]="No Windows desktop targets found. Ensure XAML files contain AttachWindow activities with Selector attributes."
    autowidth(ws9)

    # ─────────────────────────────────────────────────
    # Sheet 10 — Target Reusability Summary (GroupBy)
    # ─────────────────────────────────────────────────
    ws10 = wb.create_sheet("🎯 Target Summary")
    ws10.sheet_view.showGridLines=False

    # ── Build groupby data ──────────────────────────
    # Web: group by (App/Domain + Activity + count projects)
    web_group = defaultdict(lambda: {"projects": set(), "urls": set(),
                                     "page_titles": set(), "browsers": set(),
                                     "activity_model": set()})
    for proj in projects:
        pname = proj.results["project_info"]["name"]
        for t in proj.results["web_targets"]:
            # Key = domain + activity display name
            try:
                from urllib.parse import urlparse
                domain = urlparse(t["url"]).netloc if t["url"] else ""
            except Exception:
                domain = ""
            app_label = domain or t["page_title"] or t["browser"] or "Unknown Web App"
            key = (app_label, t["activity"])
            web_group[key]["projects"].add(pname)
            if t["url"]: web_group[key]["urls"].add(t["url"])
            if t["page_title"]: web_group[key]["page_titles"].add(t["page_title"])
            if t["browser"]: web_group[key]["browsers"].add(t["browser"])
            web_group[key]["activity_model"].add(t.get("activity_model","Classic"))

    win_group = defaultdict(lambda: {"projects": set(), "window_titles": set(),
                                     "activity_model": set()})
    for proj in projects:
        pname = proj.results["project_info"]["name"]
        for t in proj.results["window_targets"]:
            app_label = t["app_friendly"] or t["app_exe"] or "Unknown App"
            key = (app_label, t["activity"])
            win_group[key]["projects"].add(pname)
            if t["window_title"]: win_group[key]["window_titles"].add(t["window_title"])
            win_group[key]["activity_model"].add(t.get("activity_model","Classic"))

    # ── Section header ──────────────────────────────
    ws10.merge_cells("A1:H1")
    title_cell = ws10["A1"]
    title_cell.value = "Target Reusability Summary — Grouped by Application + Activity"
    title_cell.font = TITLE_F
    title_cell.alignment = Alignment(horizontal="center")
    ws10.row_dimensions[1].height = 22

    # ── Web Targets GroupBy ─────────────────────────
    ws10.append([])
    ws10.merge_cells(f"A3:H3")
    sub = ws10["A3"]
    sub.value = "🌐  WEB BROWSER TARGETS — Grouped by Domain + Activity"
    sub.font = Font(bold=True, color="FFFFFF", size=11)
    sub.fill = PatternFill("solid", fgColor="1F4E79")
    sub.alignment = Alignment(horizontal="left", vertical="center")

    web_cols = ["Application / Domain", "Activity", "Project Count",
                "Projects", "Unique URLs", "Page Titles", "Browser(s)", "Activity Model"]
    hdr(ws10, 4, web_cols)

    # Sort by Project Count desc, then total URLs desc
    web_sorted = sorted(
        web_group.items(),
        key=lambda x: (-len(x[1]["projects"]), -len(x[1]["urls"]))
    )
    for i, ((app_label, activity), data) in enumerate(web_sorted):
        proj_count = len(data["projects"])
        fill = GREEN_FILL if proj_count >= 3 else AMBER_FILL if proj_count == 2 else PatternFill()
        ws10.append([
            app_label,
            activity,
            proj_count,
            ", ".join(sorted(data["projects"])),
            "\n".join(sorted(data["urls"])),
            ", ".join(sorted(data["page_titles"])),
            ", ".join(sorted(data["browsers"])),
            ", ".join(sorted(data["activity_model"])),
        ])
        r = ws10.max_row
        for c in range(1, 9):
            cell = ws10.cell(r, c)
            if fill.patternType: cell.fill = fill
            cell.border = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws10.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws10.row_dimensions[r].height = 35

    # ── Spacer ──────────────────────────────────────
    ws10.append([])
    ws10.append([])

    # ── Window Targets GroupBy ──────────────────────
    sub2_row = ws10.max_row
    ws10.merge_cells(f"A{sub2_row}:H{sub2_row}")
    sub2 = ws10.cell(sub2_row, 1)
    sub2.value = "🖥️  WINDOWS DESKTOP TARGETS — Grouped by Application + Activity"
    sub2.font = Font(bold=True, color="FFFFFF", size=11)
    sub2.fill = PatternFill("solid", fgColor="1F3864")
    sub2.alignment = Alignment(horizontal="left", vertical="center")

    win_cols = ["Application / Process", "Activity", "Project Count",
                "Projects", "Window Titles Seen", "Activity Model", "", ""]
    hdr(ws10, ws10.max_row + 1, win_cols)

    win_sorted = sorted(
        win_group.items(),
        key=lambda x: (-len(x[1]["projects"]), -len(x[1]["window_titles"]))
    )
    for i, ((app_label, activity), data) in enumerate(win_sorted):
        proj_count = len(data["projects"])
        fill = GREEN_FILL if proj_count >= 3 else AMBER_FILL if proj_count == 2 else PatternFill()
        ws10.append([
            app_label,
            activity,
            proj_count,
            ", ".join(sorted(data["projects"])),
            "\n".join(sorted(data["window_titles"])),
            ", ".join(sorted(data["activity_model"])),
            "", "",
        ])
        r = ws10.max_row
        for c in range(1, 7):
            cell = ws10.cell(r, c)
            if fill.patternType: cell.fill = fill
            cell.border = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws10.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws10.row_dimensions[r].height = 35

    # Add legend note
    ws10.append([])
    note_row = ws10.max_row + 1
    ws10.cell(note_row, 1).value = (
        "🟢 Green = 3+ projects (high reuse candidate)    "
        "🟡 Amber = 2 projects (monitor for consolidation)    "
        "White = 1 project (project-specific)"
    )
    ws10.cell(note_row, 1).font = Font(italic=True, color="595959", size=9)
    ws10.merge_cells(f"A{note_row}:H{note_row}")

    # Fix column widths
    ws10.column_dimensions["A"].width = 32
    ws10.column_dimensions["B"].width = 30
    ws10.column_dimensions["C"].width = 14
    ws10.column_dimensions["D"].width = 40
    ws10.column_dimensions["E"].width = 45
    ws10.column_dimensions["F"].width = 35
    ws10.column_dimensions["G"].width = 20
    ws10.column_dimensions["H"].width = 16

    wb.save(filename)
    return filename


# ══════════════════════════════════════════════════════════
#  JSON Export
# ══════════════════════════════════════════════════════════
def export_json(projects, wf_matrix, act_matrix, shared_libs, output_dir: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    filename = output_dir / f"{label}_UiPath_Report_{ts}.json"

    def clean(obj):
        if isinstance(obj, set): return sorted(obj)
        if isinstance(obj, dict): return {k: clean(v) for k,v in obj.items()}
        if isinstance(obj, list): return [clean(i) for i in obj]
        return obj

    data = {
        "analyzed_on": datetime.now().isoformat(),
        "total_projects": len(projects),
        "projects": [clean(p.results) for p in projects],
        "reusability_matrix": {
            "workflow_level": wf_matrix,
            "activity_category_level": act_matrix,
            "shared_library_candidates": shared_libs,
        },
        "web_targets": [
            dict(t, project=p.results["project_info"]["name"])
            for p in projects for t in p.results["web_targets"]
        ],
        "window_targets": [
            dict(t, project=p.results["project_info"]["name"])
            for p in projects for t in p.results["window_targets"]
        ],
    }
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    return filename


# ══════════════════════════════════════════════════════════
#  CSV Export
# ══════════════════════════════════════════════════════════
def export_csv(projects, wf_matrix, act_matrix, output_dir: Path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = projects[0].results["project_info"]["name"] if len(projects)==1 else "RepoAnalysis"
    paths = []

    # Activities CSV
    rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for act in proj.results["activities"]:
            rows.append({
                "Project": pname,
                "Category": act["category"],
                "Activity Name": act["friendly_name"] or act["raw_tag"],
                "Raw Tag": act["raw_tag"],
                "Usage Count": act["count"],
                "XAML Files": ", ".join(Path(f).name for f in act["files"]),
            })
    p1 = output_dir / f"{label}_Activities_{ts}.csv"
    pd.DataFrame(rows).to_csv(p1, index=False); paths.append(p1)

    # Workflow Matrix CSV
    p2 = output_dir / f"{label}_WorkflowMatrix_{ts}.csv"
    pd.DataFrame(wf_matrix).to_csv(p2, index=False); paths.append(p2)

    # Activity Matrix CSV
    p3 = output_dir / f"{label}_ActivityMatrix_{ts}.csv"
    pd.DataFrame(act_matrix).to_csv(p3, index=False); paths.append(p3)

    # Web Targets CSV
    web_rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["web_targets"]:
            web_rows.append({"Project":pname,"URL":t["url"],"Page Title":t["page_title"],
                             "Browser":t["browser"],"Activity":t["activity"],
                             "XAML File":Path(t["xaml_file"]).name})
    p4 = output_dir / f"{label}_WebTargets_{ts}.csv"
    pd.DataFrame(web_rows).to_csv(p4, index=False); paths.append(p4)

    # Window Targets CSV
    win_rows = []
    for proj in projects:
        pname=proj.results["project_info"]["name"]
        for t in proj.results["window_targets"]:
            win_rows.append({"Project":pname,"Process (EXE)":t["app_exe"],
                             "Friendly App Name":t["app_friendly"],
                             "Window Title":t["window_title"],
                             "Activity":t["activity"],
                             "XAML File":Path(t["xaml_file"]).name})
    p5 = output_dir / f"{label}_WindowTargets_{ts}.csv"
    pd.DataFrame(win_rows).to_csv(p5, index=False); paths.append(p5)

    return paths


# ══════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        description="UiPath RPA Analyzer v2 — scans one or many UiPath projects, outputs console + Excel by default",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan a folder — auto-discovers ALL projects inside (default: console + Excel)
  python3 uipath_analyzer.py ./MyRepoOrProjectFolder

  # Specify where to save the Excel file
  python3 uipath_analyzer.py ./MyRepo --report-dir ./reports

  # Console only (skip Excel)
  python3 uipath_analyzer.py ./MyRepo --output console

  # Generate JSON or CSV instead
  python3 uipath_analyzer.py ./MyRepo --output json
  python3 uipath_analyzer.py ./MyRepo --output csv

  # Diagnose why web/window targets are not being extracted
  python3 uipath_analyzer.py ./MyProject --diagnose
        """
    )
    p.add_argument("path",         help="Project folder OR a parent folder containing multiple projects")
    p.add_argument("--output",     choices=["console","excel","json","csv"],
                   default=None,
                   help="Output format. Default: console + Excel. Use 'console' for console only.")
    p.add_argument("--report-dir", default=".", help="Directory to save the Excel file (default: current dir)")
    p.add_argument("--quiet",      action="store_true", help="Suppress console report")
    p.add_argument("--diagnose",   action="store_true",
                   help="Dump raw XML tags and attributes to help debug missing targets")
    return p.parse_args()


def main():
    banner()
    args   = parse_args()
    root   = Path(args.path)
    outdir = Path(args.report_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    # ── Auto-discover all projects under the given path ────────────
    print(f"{C.BOLD}[1/5] Scanning: {root}{C.RESET}")
    project_paths = discover_projects(root)
    if not project_paths:
        print(f"{C.RED}[ERROR] No UiPath projects (project.json) found under {root}{C.RESET}")
        sys.exit(1)
    n = len(project_paths)
    print(f"{C.GREEN}[✓] Found {n} project{'s' if n != 1 else ''}{C.RESET}")

    # ── Analyse each project ───────────────────────────────────────
    print(f"{C.BOLD}[2/5] Analysing projects...{C.RESET}")
    projects = []
    for pp in project_paths:
        analyzer = UiPathProjectAnalyzer(pp)
        if analyzer.run():
            pname = analyzer.results["project_info"]["name"]
            nxaml = len(analyzer.xaml_files)
            print(f"  {C.GREEN}✓{C.RESET}  {pname:<35} {C.GREY}({nxaml} XAML files){C.RESET}")
            projects.append(analyzer)
        else:
            print(f"  {C.YELLOW}⚠{C.RESET}  Skipped (no project.json): {pp}")

    if not projects:
        print(f"{C.RED}[ERROR] No valid projects found.{C.RESET}"); sys.exit(1)

    # ── Diagnose mode ──────────────────────────────────────────────
    if args.diagnose:
        run_diagnostics(projects)
        print(f"\n{C.CYAN}Diagnostic complete.{C.RESET}\n")
        return

    # ── Reusability matrix ─────────────────────────────────────────
    print(f"{C.BOLD}[3/5] Computing reusability matrix...{C.RESET}")
    engine    = ReusabilityEngine(projects)
    wf_matrix = engine.workflow_matrix()
    act_matrix= engine.activity_matrix()
    shared    = engine.shared_library_recommendations(wf_matrix, act_matrix)
    print(f"  {C.GREEN}✓{C.RESET}  {len(wf_matrix)} workflows scored  |  "
          f"{len(act_matrix)} categories scored  |  {len(shared)} shared lib candidates")

    # ── Determine outputs ──────────────────────────────────────────
    # Default (no --output flag): console + Excel
    # --output console  → console only
    # --output excel    → Excel only (quiet)
    # --output json/csv → that format only
    out = args.output   # None | "console" | "excel" | "json" | "csv"

    do_console = (out is None) or (out == "console")
    do_excel   = (out is None) or (out == "excel")
    do_json    = (out == "json")
    do_csv     = (out == "csv")

    # ── Console report ─────────────────────────────────────────────
    print(f"{C.BOLD}[4/5] Generating reports...{C.RESET}")
    if do_console and not args.quiet:
        for proj in projects:
            print_project_report(proj)
        print_reusability_report(wf_matrix, act_matrix, shared, len(projects))

    # ── File exports ───────────────────────────────────────────────
    print(f"{C.BOLD}[5/5] Exporting files...{C.RESET}")
    if do_excel:
        path = export_excel(projects, wf_matrix, act_matrix, shared, outdir)
        print(f"  {C.GREEN}[✓] Excel  → {path}{C.RESET}")
    if do_json:
        path = export_json(projects, wf_matrix, act_matrix, shared, outdir)
        print(f"  {C.GREEN}[✓] JSON   → {path}{C.RESET}")
    if do_csv:
        paths = export_csv(projects, wf_matrix, act_matrix, outdir)
        for path in paths:
            print(f"  {C.GREEN}[✓] CSV    → {path}{C.RESET}")
    if not do_excel and not do_json and not do_csv:
        print(f"  {C.GREY}(console only — no files exported){C.RESET}")

    print(f"\n{C.CYAN}✅  Analysis complete!{C.RESET}\n")


if __name__ == "__main__":
    main()
